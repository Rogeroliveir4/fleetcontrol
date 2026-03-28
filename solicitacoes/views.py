from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from movimentacoes.models import Movimentacao
from solicitacoes.models import SolicitacaoVeiculo
from veiculos.models import Veiculo
from motoristas.models import Motorista
from django.core.paginator import Paginator
import pandas as pd
from django.http import HttpResponse
import pandas as pd
import xlsxwriter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl
from io import BytesIO
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from .models import SolicitacaoVeiculo
from openpyxl.styles import Font, PatternFill, Alignment
from contas.models import PerfilUsuario 
from openpyxl import Workbook
from django.db.models import Q, Case, When, Value, IntegerField
from datetime import datetime, timedelta
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.contrib.auth.models import User
from django.db import transaction



# FUNÇÃO PARA NOTIFICAR GESTORES VIA EMAIL SOBRE NOVA SOLICITAÇÃO
def notificar_gestores_nova_solicitacao(request, solicitacao):

    gestores = User.objects.filter(
        perfilusuario__nivel="gestor",
        perfilusuario__contrato=solicitacao.contrato
    ).exclude(email="")

    protocol = "https" if request.is_secure() else "http"
    domain = request.get_host()

    #  URL dinâmica (melhor prática)
    url = f"{protocol}://{domain}/login/"

    for gestor in gestores:

        subject = "FleetControl - Nova solicitação aguardando aprovação"

        #  TEXTO (fallback - mantém compatibilidade)
        plain_message = f"""
        Olá {gestor.first_name or gestor.username},

        Uma nova solicitação de veículo foi criada.

        Solicitante: {solicitacao.solicitante_nome}
        Veículo: {solicitacao.veiculo}
        Destino: {solicitacao.destino}

        Acesse o sistema:
        {url}

        FleetControl
        Sistema de Gestão de pátio
        """

        #  HTML (usa seu template)
        html_message = render_to_string(
            "contas/email_nova_solicitacao.html",
            {
                "gestor_nome": gestor.first_name or gestor.username,
                "solicitacao": solicitacao,
                "url": url,
                "ano": timezone.now().year,
            }
        )

        email = EmailMultiAlternatives(
            subject=subject,
            body=plain_message,
            from_email="FleetControl <fleetcontrol.app@gmail.com>",
            to=[gestor.email],
        )

        #  Aqui ativa o HTML
        email.attach_alternative(html_message, "text/html")

        email.send(fail_silently=True)



# SOLICITAR VEÍCULO (Solicitante, Gestor ou ADM)
@login_required
def solicitar_veiculo(request, veiculo_id):

    with transaction.atomic():

        # LOCK no veículo para gravação da solicitação
        veiculo = Veiculo.objects.select_for_update().get(id=veiculo_id)

        perfil = request.user.perfilusuario

        # VERIFICA SE existe alguma solicitação ativa para esse veículo
        existe_ativa = SolicitacaoVeiculo.objects.filter(
            veiculo=veiculo,
            status__in=[
                "PENDENTE",
                "AGUARDANDO_SAIDA_PORTARIA",
                "EM_TRANSITO"
            ]
        ).exists()

        if existe_ativa:
            messages.error(
                request,
                "Este veículo já possui uma solicitação em andamento."
            )
            return redirect("lista_veiculos")

        # DEFINIR NOME
        if perfil.nome_exibicao:
            nome_solicitante = perfil.nome_exibicao
        elif perfil.nome:
            nome_solicitante = perfil.nome
        else:
            nome_solicitante = (
                request.user.get_full_name()
                or request.user.username.split("@")[0].replace(".", " ").title()
            )

        contrato = perfil.contrato or veiculo.contrato

        if perfil.nivel != "adm" and veiculo.contrato != perfil.contrato:
            messages.error(request, "Você não pode solicitar veículos de outro contrato.")
            return redirect("dashboard_solicitante")

        motoristas = Motorista.objects.filter(contrato=contrato)

        if request.method == "POST":
            motorista_id = request.POST.get("motorista")
            destino = request.POST.get("destino")
            justificativa = request.POST.get("justificativa", "")

            previsao_retorno_str = request.POST.get("previsao_retorno")
            previsao_retorno = None

            if previsao_retorno_str:
                try:
                    previsao_retorno = datetime.strptime(
                        previsao_retorno_str,
                        "%Y-%m-%dT%H:%M"
                    )
                    previsao_retorno = timezone.make_aware(previsao_retorno)
                except ValueError:
                    previsao_retorno = None

            if not motorista_id:
                messages.error(request, "Selecione um motorista.")
                return redirect("solicitar_veiculo", veiculo_id=veiculo.id)

            motorista = get_object_or_404(
                Motorista,
                id=motorista_id,
                contrato=contrato
            )

            #  REGRA: gestor/adm já cria aprovado
            if perfil.nivel == "gestor" or perfil.nivel == "adm":
                status_inicial = "AGUARDANDO_SAIDA_PORTARIA"
                data_aprovacao = timezone.now()
                gestor_responsavel = request.user
                gestor_responsavel_nome = request.user.get_full_name() or request.user.username
            else:
                status_inicial = "PENDENTE"
                data_aprovacao = None
                gestor_responsavel = None
                gestor_responsavel_nome = None

            solicitacao = SolicitacaoVeiculo.objects.create(
                origem="SISTEMA",
                veiculo=veiculo,
                motorista=motorista,
                contrato=contrato,
                id_contrato=contrato.id if contrato else None,
                destino=destino,
                justificativa=justificativa,
                previsao_retorno=previsao_retorno,
                status=status_inicial,
                data_aprovacao=data_aprovacao,
                gestor_responsavel=gestor_responsavel,
                gestor_responsavel_nome=gestor_responsavel_nome,
                solicitante=request.user,
                solicitante_nome=nome_solicitante,
                data_criacao=timezone.now()
            )

            if perfil.nivel != "gestor":
                notificar_gestores_nova_solicitacao(request, solicitacao)

            Veiculo.objects.filter(id=veiculo.id).update(status="Reservado")

            if perfil.nivel == "gestor":
                messages.success(request, "Solicitação criada e aprovada automaticamente.")
            else:
                messages.success(request, "Solicitação enviada ao gestor.")
            
            return redirect("lista_veiculos")

    # fora da transação (GET)
    return render(request, "solicitantes/solicitar.html", {
        "veiculo": veiculo,
        "motoristas": motoristas,
    })




# CANCELAR SOLICITAÇÃO (SOLICITANTE)
@login_required
def cancelar_solicitacao(request, pk):

    solicitacao = get_object_or_404(
        SolicitacaoVeiculo,
        pk=pk,
        solicitante=request.user
    )

    if solicitacao.status != "PENDENTE":
        messages.error(
            request,
            "Você só pode cancelar solicitações que ainda não foram aprovadas."
        )
        return redirect("minhas_solicitacoes")

    if request.method == "POST":

        motivo = request.POST.get("motivo_cancelamento", "").strip()

        if not motivo:
            messages.error(request, "Informe o motivo do cancelamento.")
            return redirect("minhas_solicitacoes")

        #  CANCELAMENTO
        solicitacao.status = "CANCELADA"
        solicitacao.motivo_cancelamento = motivo
        solicitacao.data_cancelamento = timezone.now()

        #  AUDITORIA (QUEM CANCELOU)
        solicitacao.cancelado_por = request.user

        perfil = getattr(request.user, "perfilusuario", None)

        if perfil:
            if perfil.nome_exibicao:
                solicitacao.cancelado_por_nome = perfil.nome_exibicao
            elif perfil.nome:
                solicitacao.cancelado_por_nome = perfil.nome
            else:
                username = request.user.username.split("@")[0]
                solicitacao.cancelado_por_nome = username.replace(".", " ").title()
        else:
            solicitacao.cancelado_por_nome = (
                request.user.get_full_name()
                or request.user.username
            )

        solicitacao.save()

        #  LIBERAÇÃO SEGURA DO VEÍCULO
        if solicitacao.veiculo_id:

            existe_outra = SolicitacaoVeiculo.objects.filter(
                veiculo_id=solicitacao.veiculo_id,
                status__in=[
                    "PENDENTE",
                    "AGUARDANDO_SAIDA_PORTARIA",
                    "EM_TRANSITO"
                ]
            ).exclude(id=solicitacao.id).exists()

            if not existe_outra:
                Veiculo.objects.filter(
                    id=solicitacao.veiculo_id
                ).update(status="Disponivel")

        messages.success(
            request,
            "Solicitação cancelada com sucesso."
        )

        return redirect("minhas_solicitacoes")

    return redirect("minhas_solicitacoes")



# DASHBOARD GESTOR - LISTAGEM COM FILTROS COMPLETOS (TELA DE APROVAÇÕES DO GESTOR)
def gestor_solicitacoes(request):
    perfil = request.user.perfilusuario
    
    # Base do queryset - filtro por contrato se não for ADM
    if perfil.nivel == "adm":
        solicitacoes = SolicitacaoVeiculo.objects.select_related(
            "motorista", 
            "veiculo",
            "gestor_responsavel",
            "gestor_reprovador",
            "contrato"
        ).all()
    else:
        solicitacoes = SolicitacaoVeiculo.objects.select_related(
            "motorista", 
            "veiculo",
            "gestor_responsavel",
            "gestor_reprovador",
            "contrato"
        ).filter(contrato=perfil.contrato)
    
    # Guardar o queryset original para estatísticas
    solicitacoes_originais = solicitacoes
    

    # CONTAGEM PARA DASHBOARD (ESTATÍSTICAS) - usar o queryset original
    total_solicitacoes = solicitacoes_originais.count()
    pendentes_count = solicitacoes_originais.filter(status="PENDENTE").count()
    # Nota: No seu template você usa "APROVADA", mas no modelo é "AGUARDANDO_SAIDA_PORTARIA"
    # Vou criar a contagem para o que aparece nos cards do template:
    aprovadas_count = solicitacoes_originais.filter(status="AGUARDANDO_SAIDA_PORTARIA").count()
    em_transito_count = solicitacoes_originais.filter(status="EM_TRANSITO").count()
    finalizadas_count = solicitacoes_originais.filter(status="FINALIZADA").count()
    reprovadas_count = solicitacoes_originais.filter(status="REPROVADA").count()
    canceladas_count = solicitacoes_originais.filter(status="CANCELADA").count()
    aguardando_saida_count = solicitacoes_originais.filter(status="AGUARDANDO_SAIDA_PORTARIA").count()
    
    # FILTROS COMPLETOS (aplicar no queryset)
    status = request.GET.get("status")
    search = request.GET.get("search")
    data_inicio = request.GET.get("inicio")
    data_fim = request.GET.get("fim")
    
    # Filtro por status
    if status and status != "TODOS":
        solicitacoes = solicitacoes.filter(status=status)
    
    # Filtro por busca (motorista, veículo, destino)
    if search:
        search = search.strip().upper()

        #  PRIORIDADE: TAG (ex: VA-132)
        if "-" in search and search.startswith("VA"):
            solicitacoes = solicitacoes.filter(
                veiculo__tag_interna__iexact=search
            )

        # 🔍 PLACA (ABC-1234)
        elif "-" in search and len(search) >= 7:
            solicitacoes = solicitacoes.filter(
                veiculo__placa__iexact=search
            )

        #  ID direto
        elif search.isdigit():
            solicitacoes = solicitacoes.filter(id=int(search))

        #  BUSCA GERAL
        else:
            solicitacoes = solicitacoes.filter(
                Q(veiculo__tag_interna__icontains=search) |   # ✅ NOVO
                Q(motorista__nome__icontains=search) |
                Q(destino__icontains=search) |
                Q(veiculo__placa__icontains=search) |
                Q(veiculo__modelo__icontains=search) |
                Q(veiculo__marca__icontains=search) |
                Q(justificativa__icontains=search)
            )
    
    # Filtro por período (data criação)
    if data_inicio:
        try:
            solicitacoes = solicitacoes.filter(data_criacao__date__gte=data_inicio)
        except:
            pass
    
    if data_fim:
        try:
            solicitacoes = solicitacoes.filter(data_criacao__date__lte=data_fim)
        except:
            pass
    
    # ORDENAÇÃO DINÂMICA
    sort = request.GET.get("sort", "-data_criacao")
    
    # Mapeamento seguro para ordenação
    sort_mapping = {
        'data_criacao': 'data_criacao',
        '-data_criacao': '-data_criacao',
        'data_aprovacao': 'data_aprovacao',
        '-data_aprovacao': '-data_aprovacao',
        'veiculo': 'veiculo__placa',
        '-veiculo': '-veiculo__placa',
        'motorista': 'motorista__nome',
        '-motorista': '-motorista__nome',
        'status': 'status',
        '-status': '-status',
        'previsao_retorno': 'previsao_retorno',
        '-previsao_retorno': '-previsao_retorno',
    }
    
    sort_field = sort_mapping.get(sort, '-data_criacao')
    solicitacoes = solicitacoes.order_by(sort_field)
    

    # ESTATÍSTICAS APÓS FILTROS
    # -----------------------------
    # Recalcular contagens após aplicar filtros
    total_filtrado = solicitacoes.count()
    
    # PAGINAÇÃO
    paginator = Paginator(solicitacoes, 15)  # 15 por página
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    
    return render(request, "solicitacoes/gestor_listar.html", {
        # Dados principais
        "solicitacoes": page_obj,
        
        # Estatísticas básicas (para os cards) - usar dados totais (não filtrados)
        "pendentes_count": pendentes_count,
        "aprovadas_count": aprovadas_count,
        "em_transito_count": em_transito_count,
        "finalizadas_count": finalizadas_count,
        "reprovadas_count": reprovadas_count,
        "canceladas_count": canceladas_count,
        "aguardando_saida_count": aguardando_saida_count,
        
        # Para o template acessar choices
        "SolicitacaoVeiculo": SolicitacaoVeiculo,
        
        # Para paginação
        "page_obj": page_obj,
        "paginator": paginator,
        
        # Para estatísticas
        "total_solicitacoes": total_solicitacoes,
        "total_filtrado": total_filtrado,
        
        # Filtros atuais (para manter no form)
        "filtro_status": status,
        "filtro_search": search,
        "filtro_inicio": data_inicio,
        "filtro_fim": data_fim,
        "sort": sort,
        
    })



# FUNÇÃO PARA CALCULAR ESTATÍSTICAS COMPLETAS DO DASHBOARD
def get_dashboard_stats(solicitacoes):
    """    Retorna estatísticas completas para o dashboard """
    from datetime import datetime, timedelta
    
    # Contagem por status
    stats = {
        'total': solicitacoes.count(),
        'por_status': {},
        'ultimos_7_dias': 0,
        'media_diaria': 0,
        'top_veiculos': [],
        'top_motoristas': [],
        'viagens_em_atraso': 0,
    }
    
    # Contagem por status
    status_counts = solicitacoes.values('status').annotate(
        count=models.Count('id')
    )
    
    for item in status_counts:
        stats['por_status'][item['status']] = item['count']
    
    # Solicitações dos últimos 7 dias
    sete_dias_atras = datetime.now() - timedelta(days=7)
    stats['ultimos_7_dias'] = solicitacoes.filter(
        data_criacao__gte=sete_dias_atras
    ).count()
    
    # Média diária (baseada nos últimos 30 dias)
    trinta_dias_atras = datetime.now() - timedelta(days=30)
    solicitacoes_30_dias = solicitacoes.filter(
        data_criacao__gte=trinta_dias_atras
    ).count()
    stats['media_diaria'] = round(solicitacoes_30_dias / 30, 1) if solicitacoes_30_dias > 0 else 0
    
    # Viagens em atraso (previsão retorno já passou e status não é FINALIZADA)
    agora = datetime.now()
    stats['viagens_em_atraso'] = solicitacoes.filter(
        previsao_retorno__lt=agora
    ).exclude(
        status__in=['FINALIZADA', 'REPROVADA']
    ).count()
    
    # Top 5 veículos mais solicitados
    stats['top_veiculos'] = solicitacoes.values(
        'veiculo__placa', 
        'veiculo__modelo',
        'veiculo__marca'
    ).annotate(
        total=models.Count('id')
    ).order_by('-total')[:5]
    
    # Top 5 motoristas mais ativos
    stats['top_motoristas'] = solicitacoes.values(
        'motorista__nome',
        'motorista__cpf'
    ).annotate(
        total=models.Count('id')
    ).order_by('-total')[:5]
    
    return stats

# ----------------------------------------------------------------------
# APROVAR SOLICITAÇÃO  (NOVO FLUXO - SEM CHECKLIST DO MOTORISTA)
# ----------------------------------------------------------------------
def aprovar_solicitacao(request, id):
    solicitacao = get_object_or_404(SolicitacaoVeiculo, id=id)
    veiculo = solicitacao.veiculo
    user = request.user

    if solicitacao.status != "PENDENTE":
        messages.warning(request, "Esta solicitação já foi analisada.")
        return redirect("gestor_solicitacoes")

    if request.method == "POST":
        observacao = request.POST.get("observacao_aprovacao", "").strip()

        solicitacao.status = "AGUARDANDO_SAIDA_PORTARIA"
        solicitacao.data_aprovacao = timezone.now()
        solicitacao.gestor_responsavel = user
        solicitacao.gestor_responsavel_nome = user.get_full_name() or user.username
        solicitacao.observacao_aprovacao = observacao or None

        solicitacao.save()

        veiculo.status = "Reservado"
        veiculo.save(update_fields=["status"])

        messages.success(request, "Solicitação aprovada com sucesso.")
        return redirect("gestor_solicitacoes")

    # fallback (caso ainda exista aprovação por GET)
    return redirect("gestor_solicitacoes")





# REPROVAR SOLICITAÇÃO (REPROVAÇÃO DO GESTOR/ADM)
def reprovar_solicitacao(request, id):
    solicitacao = get_object_or_404(SolicitacaoVeiculo, id=id)
    veiculo = solicitacao.veiculo
    user = request.user

    if solicitacao.status != "PENDENTE":
        messages.warning(request, "Esta solicitação já foi analisada.")
        return redirect("gestor_solicitacoes")

    if request.method == "POST":
        motivo = request.POST.get("motivo_reprovacao", "").strip()

        if not motivo:
            messages.error(request, "Informe o motivo da reprovação.")
            return redirect("gestor_solicitacoes")

        solicitacao.status = "REPROVADA"
        solicitacao.data_reprovacao = timezone.now()
        solicitacao.gestor_reprovador = user
        solicitacao.gestor_reprovador_nome = user.get_full_name() or user.username
        solicitacao.motivo_reprovacao = motivo

        solicitacao.save()

        veiculo.status = "Disponivel"
        veiculo.save(update_fields=["status"])

        messages.error(request, "Solicitação reprovada.")
        return redirect("gestor_solicitacoes")

    return redirect("gestor_solicitacoes")




# SOLICITAÇÕES DO USUÁRIO LOGADO () SOLICITANTE VÊ APENAS AS SUAS, GESTOR VÊ TODOS DO CONTRATO, ADM VÊ TODOS
# SOLICITAÇÕES DO USUÁRIO LOGADO (SOLICITANTE VÊ APENAS AS SUAS, GESTOR VÊ TODOS DO CONTRATO, ADM VÊ TODOS)
@login_required
def minhas_solicitacoes(request):
    perfil = request.user.perfilusuario

    # Apenas solicitante / gestor / adm
    if perfil.nivel not in ["basico", "gestor", "adm"]:
        messages.error(request, "Acesso não autorizado.")
        return redirect("dashboard_solicitante")

    # SOLICITANTE → SEMPRE pelo campo solicitante
    if perfil.nivel == "basico":
        solicitacoes = SolicitacaoVeiculo.objects.filter(
            solicitante=request.user
        )

    # GESTOR → por contrato (todos os solicitantes do contrato)
    elif perfil.nivel == "gestor":
        solicitacoes = SolicitacaoVeiculo.objects.filter(
            contrato=perfil.contrato
        )

    # ADM → tudo
    else:
        solicitacoes = SolicitacaoVeiculo.objects.all()

    # ===== FILTROS =====
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'recentes')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    # Filtro por status
    if status_filter and status_filter != 'todos':
        solicitacoes = solicitacoes.filter(status=status_filter)
    
    # ===== FILTRO POR PERÍODO =====
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            solicitacoes = solicitacoes.filter(data_criacao__date__gte=data_inicio_obj)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            solicitacoes = solicitacoes.filter(data_criacao__date__lte=data_fim_obj)
        except ValueError:
            pass
    
    # ===== BUSCA =====
    if search_query:
        from django.db.models import Q
        
        query = Q()
        
        # BUSCA POR TAG INTERNA (prioridade)
        query |= Q(veiculo__tag_interna__icontains=search_query)
        
        # Busca por ID (se for número)
        if search_query.isdigit():
            query |= Q(id=int(search_query))
        
        # Busca por outros campos de texto
        query |= Q(destino__icontains=search_query)
        query |= Q(justificativa__icontains=search_query)
        query |= Q(observacao_aprovacao__icontains=search_query)
        query |= Q(motivo_cancelamento__icontains=search_query)
        query |= Q(motivo_reprovacao__icontains=search_query)
        
        # Busca por id_contrato (se for número)
        if search_query.isdigit():
            query |= Q(id_contrato=int(search_query))
        
        # Busca por veículo
        query |= Q(veiculo__marca__icontains=search_query)
        query |= Q(veiculo__modelo__icontains=search_query)
        query |= Q(veiculo__placa__icontains=search_query)
        
        # Busca por motorista
        query |= Q(motorista__nome__icontains=search_query)
        query |= Q(motorista__cpf__icontains=search_query)
        
        # Busca por solicitante
        query |= Q(solicitante__username__icontains=search_query)
        query |= Q(solicitante__first_name__icontains=search_query)
        query |= Q(solicitante__last_name__icontains=search_query)
        query |= Q(solicitante_nome__icontains=search_query)
        
        # Busca por gestores
        query |= Q(gestor_responsavel_nome__icontains=search_query)
        query |= Q(gestor_reprovador_nome__icontains=search_query)
        query |= Q(cancelado_por_nome__icontains=search_query)
        
        solicitacoes = solicitacoes.filter(query).distinct()
    
    # Ordenação
    if sort_by == 'antigas':
        solicitacoes = solicitacoes.order_by('data_criacao')
    else:  # recentes (padrão)
        solicitacoes = solicitacoes.order_by('-data_criacao')
    
    # ===== PAGINAÇÃO =====
    from django.core.paginator import Paginator
    paginator = Paginator(solicitacoes, 10)
    page = request.GET.get('page')
    solicitacoes_page = paginator.get_page(page)
    
    context = {
        "solicitacoes": solicitacoes_page,
        "status_atual": status_filter,
        "search_atual": search_query,
        "sort_atual": sort_by,
    }
    
    return render(request, "solicitacoes/minhas.html", context)




# LISTAR SAÍDAS DA PORTARIA (PORTARIA)
@login_required
def listar_saidas_portaria(request):
    perfil = getattr(request.user, "perfilusuario", None)
    
    #  Segurança explícita
    if not perfil or perfil.nivel != "portaria":
        messages.error(request, "Acesso não autorizado.")
        return redirect("lista_movimentacoes")

    request_get = request.GET.copy()

    if not request_get.get("status"):
        request_get["status"] = "pendentes"

    request.GET = request_get
    
    # Usar função compartilhada de filtragem
    qs = filtrar_saidas_portaria(request)
    
    # Contagem para estatísticas
    total = qs.count()
    
    # Contagem por status (do queryset base, não filtrado)
    base_qs = SolicitacaoVeiculo.objects.filter(contrato=perfil.contrato)
    pendentes_count = base_qs.filter(status="AGUARDANDO_SAIDA_PORTARIA").count()
    finalizadas_count = base_qs.filter(
        status__in=["EM_TRANSITO", "FINALIZADA", "REGISTRADA_SAIDA", "CONCLUIDA"]
    ).count()
    
    # Contexto
    context = {
        "solicitacoes": qs,
        "total": total,
        "pendentes_count": pendentes_count,
        "finalizadas_count": finalizadas_count,
        "status_filtro": request.GET.get("status", ""),
        "busca": request.GET.get("busca", ""),
        "data_inicio": request.GET.get("data_inicio", ""),
        "data_fim": request.GET.get("data_fim", ""),
        "date_filter": request.GET.get("date_filter", ""),
    }
    
    return render(request, "movimentacoes/portaria_saidas.html", context)






# FILTRAR AS SAÍDAS DA PORTARIA
def filtrar_saidas_portaria(request):

    perfil = getattr(request.user, "perfilusuario", None)

    # Base query
    qs = (
        SolicitacaoVeiculo.objects
        .select_related("veiculo", "motorista", "contrato")
        .order_by("-data_aprovacao")
    )

    # Filtrar por contrato
    if perfil and perfil.contrato:
        qs = qs.filter(contrato=perfil.contrato)

    # Helper para normalizar valores
    def normalizar(valor):
        if valor in [None, "", "None", "null", "undefined"]:
            return None
        return valor

    # Obter parâmetros
    status = normalizar(request.GET.get("status"))
    date_filter = normalizar(request.GET.get("date_filter"))
    data_inicio = normalizar(request.GET.get("data_inicio"))
    data_fim = normalizar(request.GET.get("data_fim"))
    busca = normalizar(request.GET.get("busca"))

    # FILTRO POR STATUS
    if status == "pendentes":
        qs = qs.filter(status="AGUARDANDO_SAIDA_PORTARIA")
    elif status == "finalizadas":
        qs = qs.filter(
            status__in=["EM_TRANSITO", "FINALIZADA", "REGISTRADA_SAIDA", "CONCLUIDA"]
        )

    hoje = timezone.now().date()

    # CORREÇÃO: FILTRO POR PERÍODO - SEMPRE USAR data_criacao OU data_saida
    if date_filter in ["today", "week", "month"]:
        # Para todas as solicitações, usar data_criacao como referência
        # Exceto se já tiver data_saida, então usar data_saida
        
        if date_filter == "today":
            qs = qs.filter(
                Q(data_saida__date=hoje) | 
                Q(data_saida__isnull=True, data_criacao__date=hoje)
            )
        
        elif date_filter == "week":
            semana_passada = hoje - timedelta(days=7)
            qs = qs.filter(
                Q(data_saida__date__gte=semana_passada) | 
                Q(data_saida__isnull=True, data_criacao__date__gte=semana_passada)
            )
        
        elif date_filter == "month":
            qs = qs.filter(
                Q(data_saida__date__month=hoje.month) | 
                Q(data_saida__date__year=hoje.year, data_saida__isnull=False) |
                Q(data_saida__isnull=True, data_criacao__date__month=hoje.month, data_criacao__date__year=hoje.year)
            )
    
    elif data_inicio or data_fim:
        # CORREÇÃO: Sempre usar data_saida se existir, senão data_criacao
        if data_inicio:
            try:
                data_inicio_obj = datetime.strptime(data_inicio, "%Y-%m-%d").date()
                qs = qs.filter(
                    Q(data_saida__date__gte=data_inicio_obj) | 
                    Q(data_saida__isnull=True, data_criacao__date__gte=data_inicio_obj)
                )
            except ValueError:
                pass
        
        if data_fim:
            try:
                data_fim_obj = datetime.strptime(data_fim, "%Y-%m-%d").date()
                qs = qs.filter(
                    Q(data_saida__date__lte=data_fim_obj) | 
                    Q(data_saida__isnull=True, data_criacao__date__lte=data_fim_obj)
                )
            except ValueError:
                pass

    # FILTRO POR BUSCA
    if busca:
        qs = qs.filter(
            Q(destino__icontains=busca) |
            Q(motorista__nome__icontains=busca) |
            Q(veiculo__placa__icontains=busca) |
            Q(veiculo__modelo__icontains=busca) |
            Q(veiculo__marca__icontains=busca) |
            Q(veiculo__tag_interna__icontains=busca) |
            Q(solicitante_nome__icontains=busca) |
            Q(justificativa__icontains=busca)
        )

    return qs






# FUNÇÃO AUXILIAR PARA NORMALIZAR VALORES DE FILTRO (TRATAR CASOS DE "None", "", etc)
def normalizar(valor):
    if valor in [None, "", "None", "null", "undefined"]:
        return None
    return valor


# VISUALIZAR DETALHES DA SOLICITAÇÃO (QUALQUER USUÁRIO COM ACESSO)
@login_required
def visualizar_solicitacao(request, pk):
    solicitacao = get_object_or_404(SolicitacaoVeiculo, pk=pk)
    return render(
        request,
        "solicitacoes/detalhe.html",
        {"solicitacao": solicitacao}
    )

# DETALHES DA SOLICITAÇÃO (COM INFORMAÇÕES COMPLETAS PARA GESTORES/ADM)
@login_required
def solicitacao_detalhe(request, pk):
    solicitacao = get_object_or_404(SolicitacaoVeiculo, pk=pk)

    return render(
        request,
        "solicitacoes/detalhe.html",
        {
            "solicitacao": solicitacao
        }
    )


# EXPORTAR SAÍDAS PORTARIA PARA EXCEL
def exportar_saidas_portaria_excel(request):
    """
    Exporta as saídas da portaria para Excel com filtros aplicados
    """
    from solicitacoes.models import SolicitacaoVeiculo
    
    # Aplicar os mesmos filtros da view principal
    solicitacoes = SolicitacaoVeiculo.objects.all().order_by('-data_criacao')
    
    # Aplicar filtro de status
    status_filtro = request.GET.get('status', '')
    if status_filtro == 'pendentes':
        solicitacoes = solicitacoes.filter(
            status='AGUARDANDO_SAIDA_PORTARIA'
        )
    elif status_filtro == 'finalizadas':
        solicitacoes = solicitacoes.filter(
            status__in=['EM_TRANSITO', 'FINALIZADA']
        )
    
    # Aplicar filtro de data
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    if data_inicio:
        try:
            data_inicio = timezone.datetime.strptime(data_inicio, '%Y-%m-%d').date()
            solicitacoes = solicitacoes.filter(data_criacao__date__gte=data_inicio)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim = timezone.datetime.strptime(data_fim, '%Y-%m-%d').date()
            solicitacoes = solicitacoes.filter(data_criacao__date__lte=data_fim)
        except ValueError:
            pass
    
    # Aplicar filtro de busca
    busca = request.GET.get('busca', '')
    if busca:
        solicitacoes = solicitacoes.filter(
            Q(motorista__nome__icontains=busca) |
            Q(veiculo__placa__icontains=busca) |
            Q(veiculo__tag_interna__icontains=busca) |
            Q(destino__icontains=busca)
        )
    
    # Criar workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Saídas Portaria"
    
    # Cabeçalhos - AJUSTADO para refletir o que realmente existe no modelo
    headers = [
        'ID', 
        'Status', 
        'Data Solicitação', 
        'Motorista', 
        'CPF', 
        'Veículo', 
        'Placa', 
        'Tag Interna', 
        'Destino',
        'Data Saída Real', 
        'Data Retorno Real', 
        'Previsão Retorno',
        'Justificativa',
        'Contrato',
        'Solicitante',
        'Gestor Responsável'
    ]
    
    # Estilos
    header_fill = PatternFill(start_color='00594C', end_color='00594C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escrever cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Função auxiliar para formatar datas
    def formatar_data(valor):
        if isinstance(valor, datetime):
            return valor.strftime("%d/%m/%Y %H:%M")
        return
    
    # Escrever dados
    for row_num, solicitacao in enumerate(solicitacoes, 2):
        # Acessar dados de forma segura
        try:
            motorista_nome = solicitacao.motorista.nome if solicitacao.motorista else ''
            motorista_cpf = solicitacao.motorista.cpf if solicitacao.motorista else ''
        except:
            motorista_nome = ''
            motorista_cpf = ''
        
        try:
            veiculo_modelo = f"{solicitacao.veiculo.marca} {solicitacao.veiculo.modelo}" if solicitacao.veiculo else ''
            veiculo_placa = solicitacao.veiculo.placa if solicitacao.veiculo else ''
            veiculo_tag = solicitacao.veiculo.tag_interna if solicitacao.veiculo else ''
        except:
            veiculo_modelo = ''
            veiculo_placa = ''
            veiculo_tag = ''
        
        # Dados da linha - APENAS CAMPOS QUE EXISTEM NO MODELO
        data = [
            solicitacao.id,
            solicitacao.get_status_display(),
            formatar_data(solicitacao.data_criacao),
            motorista_nome,
            motorista_cpf,
            veiculo_modelo,
            veiculo_placa,
            veiculo_tag,
            solicitacao.destino or '',
            formatar_data(solicitacao.data_saida),
            formatar_data(solicitacao.data_retorno),
            formatar_data(solicitacao.previsao_retorno),
            solicitacao.justificativa or '',
            str(solicitacao.contrato) if solicitacao.contrato else '',
            solicitacao.solicitante_nome or (solicitacao.solicitante.username if solicitacao.solicitante else ''),
            solicitacao.gestor_responsavel_nome or (solicitacao.gestor_responsavel.username if solicitacao.gestor_responsavel else ''),
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value or ''))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar resposta
    filename = f'saidas_portaria_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
    
    wb.save(response)
    return response



#DETALHES DA SOLICITAÇÃO
@login_required
def detalhes_solicitacao(request, solicitacao_id):
    """
    View para exibir detalhes completos de uma solicitação de veículo
    """
    # Buscar a solicitação ou retornar 404
    solicitacao = get_object_or_404(SolicitacaoVeiculo, id=solicitacao_id)
    
    # Preparar dados adicionais
    context = {
        'solicitacao': solicitacao,
        'tem_checklist_saida': hasattr(solicitacao, 'checklistsaida'),
        'tem_checklist_retorno': hasattr(solicitacao, 'checklistretorno'),
        'status_info': get_status_info(solicitacao.status),
        'tem_historico': solicitacao.historico.exists() if hasattr(solicitacao, 'historico') else False,
    }
    
    # Adicionar checklists se existirem
    if context['tem_checklist_saida']:
        context['checklist_saida'] = solicitacao.checklistsaida
    
    if context['tem_checklist_retorno']:
        context['checklist_retorno'] = solicitacao.checklistretorno
    
    # Adicionar histórico se existir
    if context['tem_historico']:
        context['historico'] = solicitacao.historico.all().order_by('-data_alteracao')
    
    return render(request, 'solicitacoes/detalhes_solicitacao.html', context)


def get_status_info(status):
    """
    Retorna informações sobre o status (cor, ícone, descrição)
    """
    status_map = {
        'PENDENTE': {
            'cor': 'yellow',
            'icone': 'fa-clock',
            'descricao': 'Aguardando aprovação',
            'bg_class': 'bg-yellow-50 dark:bg-yellow-900/30',
            'text_class': 'text-yellow-800 dark:text-yellow-300',
            'border_class': 'border-yellow-200 dark:border-yellow-800',
        },
        'AGUARDANDO_CHECKLIST': {
            'cor': 'blue',
            'icone': 'fa-clipboard-check',
            'descricao': 'Aguardando checklist de saída',
            'bg_class': 'bg-blue-50 dark:bg-blue-900/30',
            'text_class': 'text-blue-800 dark:text-blue-300',
            'border_class': 'border-blue-200 dark:border-blue-800',
        },
        'AGUARDANDO_SAIDA_PORTARIA': {
            'cor': 'indigo',
            'icone': 'fa-door-open',
            'descricao': 'Aguardando liberação da portaria',
            'bg_class': 'bg-indigo-50 dark:bg-indigo-900/30',
            'text_class': 'text-indigo-800 dark:text-indigo-300',
            'border_class': 'border-indigo-200 dark:border-indigo-800',
        },
        'EM_TRANSITO': {
            'cor': 'purple',
            'icone': 'fa-car-side',
            'descricao': 'Veículo em trânsito',
            'bg_class': 'bg-purple-50 dark:bg-purple-900/30',
            'text_class': 'text-purple-800 dark:text-purple-300',
            'border_class': 'border-purple-200 dark:border-purple-800',
        },
        'AGUARDANDO_CHECKLIST_RETORNO': {
            'cor': 'orange',
            'icone': 'fa-clipboard-list',
            'descricao': 'Aguardando checklist de retorno',
            'bg_class': 'bg-orange-50 dark:bg-orange-900/30',
            'text_class': 'text-orange-800 dark:text-orange-300',
            'border_class': 'border-orange-200 dark:border-orange-800',
        },
        'FINALIZADA': {
            'cor': 'green',
            'icone': 'fa-flag-checkered',
            'descricao': 'Solicitação finalizada',
            'bg_class': 'bg-green-50 dark:bg-green-900/30',
            'text_class': 'text-green-800 dark:text-green-300',
            'border_class': 'border-green-200 dark:border-green-800',
        },
        'REPROVADA': {
            'cor': 'red',
            'icone': 'fa-ban',
            'descricao': 'Solicitação reprovada',
            'bg_class': 'bg-red-50 dark:bg-red-900/30',
            'text_class': 'text-red-800 dark:text-red-300',
            'border_class': 'border-red-200 dark:border-red-800',
        },
    }
    
    return status_map.get(status, {
        'cor': 'gray',
        'icone': 'fa-question-circle',
        'descricao': 'Status desconhecido',
        'bg_class': 'bg-gray-50 dark:bg-gray-900/30',
        'text_class': 'text-gray-800 dark:text-gray-300',
        'border_class': 'border-gray-200 dark:border-gray-800',
    })


# EXPORTAR SOLICITAÇÕES PARA EXCEL    
@login_required
def exportar_excel_solicitacoes(request):
        perfil = request.user.perfilusuario

        # Importar datetime se não estiver importado
        from datetime import datetime

        # Filtros recebidos via GET
        status_filter = request.GET.get("status", "").strip()
        search_query = request.GET.get("search", "").strip()
        motorista_filter = request.GET.get("motorista", "").strip()
        veiculo_filter = request.GET.get("veiculo", "").strip()
        data_inicio = request.GET.get("data_inicio", "").strip()
        data_fim = request.GET.get("data_fim", "").strip()

        # Base - exportar solicitações
        solicitacoes = SolicitacaoVeiculo.objects.all().order_by("-data_criacao")

        # Permissão por contrato (não ADM)
        if perfil.nivel != "adm" and perfil.contrato:
            solicitacoes = solicitacoes.filter(contrato=perfil.contrato)

        # Filtro status
        if status_filter:
            solicitacoes = solicitacoes.filter(status=status_filter)

        # Filtro de busca
        if search_query:
            solicitacoes = solicitacoes.filter(
                Q(destino__icontains=search_query) |
                Q(justificativa__icontains=search_query) |
                Q(motorista__nome__icontains=search_query) |
                Q(veiculo__placa__icontains=search_query) |
                Q(veiculo__modelo__icontains=search_query) |  # Adicionado
                Q(veiculo__marca__icontains=search_query)     # Adicionado
            )

        # Filtro motorista
        if motorista_filter:
            solicitacoes = solicitacoes.filter(motorista__id=motorista_filter)

        # Filtro veículo
        if veiculo_filter:
            solicitacoes = solicitacoes.filter(veiculo__id=veiculo_filter)

        # Filtro por data
        if data_inicio:
            try:
                data_inicio_obj = datetime.strptime(data_inicio, "%Y-%m-%d")
                solicitacoes = solicitacoes.filter(data_criacao__date__gte=data_inicio_obj)
            except ValueError:
                pass
        
        if data_fim:
            try:
                data_fim_obj = datetime.strptime(data_fim, "%Y-%m-%d")
                solicitacoes = solicitacoes.filter(data_criacao__date__lte=data_fim_obj)
            except ValueError:
                pass

        # Criar workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Solicitações"

        # Estilo cabeçalho
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="00594C", end_color="00594C", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        # CORREÇÃO: Usar nomes de campos que realmente existem no modelo
        headers = [
            "ID",
            "Data Criação",
            "Motorista",
            "CPF Motorista",
            "Veículo",
            "Placa",
            "Destino",
            "Justificativa",
            "Status",
            "Data Saída",  # Alterado de "Data Saída Prevista"
            "Previsão Retorno",  # Alterado de "Data Retorno Prevista"
            "Data Retorno",  # Adicionado campo existente
            "Gestor Responsável",
            "Data Aprovação",
            "Gestor Reprovador",
            "Data Reprovação",
            "Contrato",
        ]

        # Escrever cabeçalhos
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            sheet.column_dimensions[cell.column_letter].width = max(len(header) + 2, 14)

        # Escrever dados das SOLICITAÇÕES - CORRIGIDO
        for row, s in enumerate(solicitacoes, start=2):
            sheet.cell(row=row, column=1, value=s.id)
            sheet.cell(row=row, column=2, value=s.data_criacao.strftime("%d/%m/%Y %H:%M") if s.data_criacao else "")
            sheet.cell(row=row, column=3, value=s.motorista.nome if s.motorista else "")
            sheet.cell(row=row, column=4, value=s.motorista.cpf if s.motorista else "")
            sheet.cell(row=row, column=5, value=f"{s.veiculo.marca} {s.veiculo.modelo}" if s.veiculo else "")
            sheet.cell(row=row, column=6, value=s.veiculo.placa if s.veiculo else "")
            sheet.cell(row=row, column=7, value=s.destino)
            sheet.cell(row=row, column=8, value=s.justificativa or "")
            
            # Status com display name
            status_display = dict(SolicitacaoVeiculo.STATUS_CHOICES).get(s.status, s.status)
            sheet.cell(row=row, column=9, value=status_display)
            
            # CORREÇÃO: Usar campo data_saida (que existe) em vez de data_saida_prevista
            sheet.cell(row=row, column=10, value=s.data_saida.strftime("%d/%m/%Y %H:%M") if s.data_saida else "")
            
            # CORREÇÃO: Usar campo previsao_retorno (que existe) em vez de data_retorno_prevista
            sheet.cell(row=row, column=11, value=s.previsao_retorno.strftime("%d/%m/%Y %H:%M") if s.previsao_retorno else "")
            
            # Adicionar campo data_retorno que existe no modelo
            sheet.cell(row=row, column=12, value=s.data_retorno.strftime("%d/%m/%Y %H:%M") if s.data_retorno else "")
            
            sheet.cell(row=row, column=13, value=s.gestor_responsavel.get_full_name() if s.gestor_responsavel else "")
            sheet.cell(row=row, column=14, value=s.data_aprovacao.strftime("%d/%m/%Y %H:%M") if s.data_aprovacao else "")
            
            # Adicionar nome do gestor reprovador
            sheet.cell(row=row, column=15, value=s.gestor_reprovador_nome or "")
            
            sheet.cell(row=row, column=16, value=s.data_reprovacao.strftime("%d/%m/%Y %H:%M") if s.data_reprovacao else "")
            sheet.cell(row=row, column=17, value=str(s.contrato) if s.contrato else "")

        # Ajustar largura das colunas automaticamente
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limite de 50 caracteres
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Criar buffer em memória
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        # Adicionar timestamp ao nome do arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"solicitacoes_{timestamp}.xlsx"
        
        # Responder com arquivo
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# EDITAR SOLICITAÇÃO (APENAS SE ESTIVER PENDENTE E FOR O SOLICITANTE)
@login_required
def editar_solicitacao(request, pk):
    solicitacao = get_object_or_404(
        SolicitacaoVeiculo,
        pk=pk,
        solicitante=request.user
    )

    if solicitacao.status != "PENDENTE":
        messages.error(
            request,
            "Esta solicitação não pode mais ser editada."
        )
        return redirect("minhas_solicitacoes")

    contrato = solicitacao.contrato
    motoristas = Motorista.objects.filter(contrato=contrato)

    if request.method == "POST":
        motorista_id = request.POST.get("motorista")
        destino = request.POST.get("destino")
        justificativa = request.POST.get("justificativa", "")

        # AJustar previsão de retorno (que agora é um campo datetime, não date)
        previsao_retorno_str = request.POST.get("previsao_retorno")
        previsao_retorno = None

        if previsao_retorno_str:
            try:
                previsao_retorno = datetime.strptime(
                    previsao_retorno_str,
                    "%Y-%m-%dT%H:%M"
                )
                previsao_retorno = timezone.make_aware(previsao_retorno)
            except ValueError:
                previsao_retorno = None

        if not motorista_id:
            messages.error(request, "Selecione um motorista.")
            return redirect("editar_solicitacao", pk=pk)

        motorista = get_object_or_404(
            Motorista,
            id=motorista_id,
            contrato=contrato
        )

        solicitacao.motorista = motorista
        solicitacao.destino = destino
        solicitacao.justificativa = justificativa
        solicitacao.previsao_retorno = previsao_retorno
        solicitacao.save()

        messages.success(request, "Solicitação atualizada com sucesso.")
        return redirect("minhas_solicitacoes")

    return render(request, "solicitantes/solicitar.html", {
        "veiculo": solicitacao.veiculo,
        "motoristas": motoristas,
        "solicitacao": solicitacao,
        "modo_edicao": True
    })



# EXPORTAR MINHAS SOLICITAÇÕES PARA EXCEL (USUÁRIOS PERFIL)
@login_required
def exportar_minhas_solicitacoes(request):
    """Exporta as solicitações do usuário logado para Excel com filtros aplicados"""
    
    perfil = request.user.perfilusuario
    from datetime import datetime
    
    # ===== BUSCAR SOLICITAÇÕES DO USUÁRIO =====
    if perfil.nivel == "basico":
        solicitacoes = SolicitacaoVeiculo.objects.filter(
            solicitante=request.user
        )
    elif perfil.nivel == "gestor":
        solicitacoes = SolicitacaoVeiculo.objects.filter(
            contrato=perfil.contrato
        )
    else:  # adm
        solicitacoes = SolicitacaoVeiculo.objects.all()
    
    # ===== APLICAR OS MESMOS FILTROS DA VIEW MINHAS_SOLICITACOES =====
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'recentes')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    # Filtro por status
    if status_filter and status_filter != 'todos':
        solicitacoes = solicitacoes.filter(status=status_filter)
    
    # Filtro por período
    if data_inicio:
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            solicitacoes = solicitacoes.filter(data_criacao__date__gte=data_inicio_obj)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            solicitacoes = solicitacoes.filter(data_criacao__date__lte=data_fim_obj)
        except ValueError:
            pass
    
    # Filtro por busca
    if search_query:
        from django.db.models import Q
        query = Q()
        query |= Q(veiculo__tag_interna__icontains=search_query)
        query |= Q(motorista__nome__icontains=search_query)
        query |= Q(destino__icontains=search_query)
        query |= Q(id__icontains=search_query) if search_query.isdigit() else Q()
        query |= Q(justificativa__icontains=search_query)
        query |= Q(solicitante_nome__icontains=search_query)
        solicitacoes = solicitacoes.filter(query).distinct()
    
    # Ordenação
    if sort_by == 'antigas':
        solicitacoes = solicitacoes.order_by('data_criacao')
    else:
        solicitacoes = solicitacoes.order_by('-data_criacao')
    
    # ===== CRIAR WORKBOOK =====
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Minhas Solicitações"
    
    # Estilo cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00594C", end_color="00594C", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    headers = [
        "ID", "Data Criação", "Veículo", "Placa", "Tag", "Motorista",
        "Destino", "Justificativa", "Status", "Previsão Retorno",
        "Data Saída", "Data Retorno", "Gestor Responsável", "Data Aprovação"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    for row, s in enumerate(solicitacoes, 2):
        sheet.cell(row=row, column=1, value=s.id)
        sheet.cell(row=row, column=2, value=s.data_criacao.strftime("%d/%m/%Y %H:%M") if s.data_criacao else "")
        sheet.cell(row=row, column=3, value=f"{s.veiculo.marca} {s.veiculo.modelo}" if s.veiculo else "")
        sheet.cell(row=row, column=4, value=s.veiculo.placa if s.veiculo else "")
        sheet.cell(row=row, column=5, value=s.veiculo.tag_interna if s.veiculo else "")
        sheet.cell(row=row, column=6, value=s.motorista.nome if s.motorista else "")
        sheet.cell(row=row, column=7, value=s.destino)
        sheet.cell(row=row, column=8, value=s.justificativa or "")
        
        status_display = dict(SolicitacaoVeiculo.STATUS_CHOICES).get(s.status, s.status)
        sheet.cell(row=row, column=9, value=status_display)
        sheet.cell(row=row, column=10, value=s.previsao_retorno.strftime("%d/%m/%Y %H:%M") if s.previsao_retorno else "")
        sheet.cell(row=row, column=11, value=s.data_saida.strftime("%d/%m/%Y %H:%M") if s.data_saida else "")
        sheet.cell(row=row, column=12, value=s.data_retorno.strftime("%d/%m/%Y %H:%M") if s.data_retorno else "")
        sheet.cell(row=row, column=13, value=s.gestor_responsavel_nome or "")
        sheet.cell(row=row, column=14, value=s.data_aprovacao.strftime("%d/%m/%Y %H:%M") if s.data_aprovacao else "")
    
    # Ajustar largura das colunas
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"minhas_solicitacoes_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response