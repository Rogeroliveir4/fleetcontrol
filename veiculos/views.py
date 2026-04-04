from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from veiculos.models import Veiculo
from contratos.models import Contrato
from contas.models import PerfilUsuario
from motoristas.models import Motorista
from solicitacoes.models import SolicitacaoVeiculo
from django.db.models import Exists, OuterRef
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from .models import Veiculo, Contrato
from .forms import VeiculoForm
from django.urls import reverse
from datetime import datetime, timedelta
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from veiculos.models import  HistoricoKM
from openpyxl import Workbook
import traceback
import re
from datetime import date
from django.http import HttpResponse
from openpyxl import Workbook
from collections import defaultdict
from .utils import validar_placa, formatar_placa, obter_tipo_placa




# LISTAGEM DE VEÍCULOS
# ----------------------------------------------------------------------
def lista_veiculos(request):

    perfil, created = PerfilUsuario.objects.get_or_create(
        user=request.user,
        defaults={"nivel": "basico"}
    )

    # Filtros GET
    status_filter = request.GET.get('status', '').strip()
    search_query = request.GET.get('search', '').strip()
    categoria_filter = request.GET.get('categoria', '').strip()
    localizacao_filter = request.GET.get('localizacao', '').strip()

    # Base inicial
    veiculos_list = Veiculo.objects.all().order_by("-ativo", "placa")

    #  Filtrar por contrato (gestor/motorista)
    if perfil.nivel != "adm" and perfil.contrato_id:
        veiculos_list = veiculos_list.filter(contrato_id=perfil.contrato_id)

    #  Motorista só vê veículos disponíveis
    if perfil.nivel == "basico":
        veiculos_list = veiculos_list.all()

    #  ANOTAR se o veículo possui solicitação pendente
    veiculos_list = veiculos_list.annotate(
        solicitacao_pendente=Exists(
            SolicitacaoVeiculo.objects.filter(
                veiculo_id=OuterRef("pk"),
                status__in=[
                    "PENDENTE",
                    "AGUARDANDO_CHECKLIST_SAIDA",
                    "AGUARDANDO_SAIDA_PORTARIA",
                    "EM_ANDAMENTO",
                    "AGUARDANDO_CHECKLIST_RETORNO",
                ]
            )
        )
    )

    #  Solicitação pendente DO USUÁRIO (motorista, gestor, adm)
    solicitacao_do_motorista = None
    solicitacao_ativa = False

    motorista = Motorista.objects.filter(user=request.user).first()
    
    if motorista:
        solicitacao_do_motorista = SolicitacaoVeiculo.objects.filter(
            motorista=motorista,
            status__in=[
            "PENDENTE",
            "AGUARDANDO_CHECKLIST_SAIDA",
            "AGUARDANDO_SAIDA_PORTARIA",
            "EM_ANDAMENTO",
            "AGUARDANDO_CHECKLIST_RETORNO"
        ]
        ).first()

        # Se existe solicitação ativa, bloquear solicitação em todos os perfis
        solicitacao_ativa = solicitacao_do_motorista is not None

    #  Filtragem por status
    if status_filter:
        mapa_status = {
            "disponivel": "Disponivel",
            "manutencao": "Manutencao",
        }

        if status_filter == "inativo":
            veiculos_list = veiculos_list.filter(ativo=False)

        elif status_filter in mapa_status:
            veiculos_list = veiculos_list.filter(
                ativo=True, status=mapa_status[status_filter]
            )

        elif status_filter == "indisponivel":
            veiculos_list = veiculos_list.filter(ativo=True).exclude(
                status="Disponivel"
            ).exclude(status="Manutencao")

        elif status_filter in dict(Veiculo.STATUS_CHOICES):
            veiculos_list = veiculos_list.filter(
                ativo=True, status=status_filter
            )

    else:
        veiculos_list = veiculos_list.filter(ativo=True)

    # ---------------------------------------------------------------
    #  Busca geral
    # ---------------------------------------------------------------
    if search_query:
        veiculos_list = veiculos_list.filter(
            Q(placa__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(marca__icontains=search_query) |
            Q(renavam__icontains=search_query) |
            Q(tag_interna__icontains=search_query) |
            Q(tag_cliente__icontains=search_query) |
            Q(contrato__nome__icontains=search_query)
        )

    # Filtro categoria
    if categoria_filter:
        veiculos_list = veiculos_list.filter(categoria=categoria_filter)

    # Filtro localização
    if localizacao_filter and hasattr(Veiculo, "localizacao"):
        veiculos_list = veiculos_list.filter(localizacao__icontains=localizacao_filter)

    # ---------------------------------------------------------------
    #  CONTAGENS AJUSTADAS
    # ---------------------------------------------------------------
    if perfil.nivel == "adm":
        base_qs = Veiculo.objects.all()
    else:
        base_qs = Veiculo.objects.filter(contrato_id=perfil.contrato_id)

    total_geral = base_qs.count()
    total_ativos = base_qs.filter(ativo=True).count()
    total_inativos = base_qs.filter(ativo=False).count()
    total_disponivel = base_qs.filter(status="Disponivel", ativo=True).count()
    total_manutencao = base_qs.filter(status="Manutencao", ativo=True).count()
    total_indisponivel = base_qs.filter(ativo=True).exclude(
        status="Disponivel"
    ).exclude(status="Manutencao").count()

    # ---------------------------------------------------------------
    #  Paginação
    # ---------------------------------------------------------------
    page = request.GET.get("page", 1)
    paginator = Paginator(veiculos_list, 20)  # 20 veículos por página
    veiculos = paginator.get_page(page)

    # ---------------------------------------------------------------
    #  AJAX RESPONSE
    # ---------------------------------------------------------------
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({
            "html_cards": render_to_string("veiculos/partials/cards.html", {
                "veiculos": veiculos,
                "solicitacao_ativa": solicitacao_ativa,
                "solicitacao_do_motorista": solicitacao_do_motorista,
            }),
            "paginacao_mobile": render_to_string("veiculos/partials/paginacao_mobile.html", {"veiculos": veiculos}),
            "paginacao_desktop": render_to_string("veiculos/partials/paginacao_desktop.html", {"veiculos": veiculos}),
            "status": status_filter,
            "total_geral": total_geral,
            "total_ativos": total_ativos,
            "total_inativos": total_inativos,
            "total_disponivel": total_disponivel,
            "total_manutencao": total_manutencao,
            "total_indisponivel": total_indisponivel,
        })

    # ---------------------------------------------------------------
    #  RENDER FINAL
    # ---------------------------------------------------------------
    return render(request, "veiculos/lista.html", {
        "veiculos": veiculos,
        "solicitacao_ativa": solicitacao_ativa,
        "solicitacao_do_motorista": solicitacao_do_motorista,
        "status": status_filter,
        "search": search_query,
        "categoria": categoria_filter,
        "localizacao": localizacao_filter,
        "total_geral": total_geral,
        "total_ativos": total_ativos,
        "total_inativos": total_inativos,
        "total_disponivel": total_disponivel,
        "total_manutencao": total_manutencao,
        "total_indisponivel": total_indisponivel,
    })

# FUNÇÕES DE VALIDAÇÃO AJAX PARA CRIAÇÃO/EDIÇÃO DE VEÍCULOS
@require_GET
def check_placa(request):
    """Verifica se uma placa já existe"""
    placa = request.GET.get('placa', '').upper().strip()
    exclude_id = request.GET.get('exclude_id', None)
    
    if not placa:
        return JsonResponse({'exists': False})
    
    # Remove formatação para busca
    placa_clean = placa.replace('-', '').replace(' ', '')
    
    # Query inicial
    qs = Veiculo.objects.filter(
        placa__iregex=r'^{}$|^{}$'.format(placa, placa_clean)
    )
    
    # Exclui o próprio veículo se for edição
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    
    exists = qs.exists()
    return JsonResponse({'exists': exists})

@require_GET
def check_tag_interna(request):
    """Verifica se uma tag interna já existe"""
    tag_interna = request.GET.get('tag_interna', '').strip()
    exclude_id = request.GET.get('exclude_id', None)
    
    if not tag_interna:
        return JsonResponse({'exists': False})
    
    # Query inicial
    qs = Veiculo.objects.filter(tag_interna__iexact=tag_interna)
    
    # Exclui o próprio veículo se for edição
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    
    exists = qs.exists()
    return JsonResponse({'exists': exists})


# ----------------------------------------------------------------------
# CRIAR VEÍCULO (com validações completas)
# ----------------------------------------------------------------------
def criar_veiculo(request):

    if request.method == "POST":

        # Normalização dos campos
        placa = request.POST.get("placa", "").upper().strip()
        tag_interna = (request.POST.get("tag_interna") or "").strip()
        #tag_cliente = (request.POST.get("tag_cliente") or "").strip()
        renavam = (request.POST.get("renavam") or "").strip()
        apolice = (request.POST.get("apolice_numero") or "").strip()

        # Função auxiliar para retornar o form com contratos
        def render_form_com_erro(msg):
            messages.error(request, msg)
            return render(request, "veiculos/novo_veiculo.html", {
                "contratos": Contrato.objects.all().order_by("id")
            })

        # -----------------------
        #  1. Validar PLACA
        # -----------------------
        if not placa:
            return render_form_com_erro("Placa é obrigatória!")

        if Veiculo.objects.filter(placa=placa).exists():
            return render_form_com_erro("Já existe outro veículo com essa placa!")

        # -----------------------
        #  2. Validar TAG INTERNA
        # -----------------------
        if tag_interna and Veiculo.objects.filter(tag_interna=tag_interna).exists():
            return render_form_com_erro("Já existe outro veículo com esta TAG interna!")

        # -----------------------
        #  3. Validar TAG CLIENTE
        # -----------------------
        #if tag_cliente and Veiculo.objects.filter(tag_cliente=tag_cliente).exists():
            #return render_form_com_erro("Já existe outro veículo com esta TAG do cliente!")

        # -----------------------
        #  4. Validar RENAVAM
        # -----------------------
        if renavam and Veiculo.objects.filter(renavam=renavam).exists():
            return render_form_com_erro("Já existe outro veículo com este Renavam!")

        # -----------------------
        #  5. Validar APÓLICE
        # -----------------------
        if apolice and Veiculo.objects.filter(apolice_numero=apolice).exists():
            return render_form_com_erro("Número de apólice já cadastrado para outro veículo!")

        # -----------------------
        # CONTRATO
        # -----------------------
        contrato_id = request.POST.get("id_contrato")
        contrato_obj = Contrato.objects.filter(id=contrato_id).first() if contrato_id else None

        # -----------------------
        # Salvar veículo
        # -----------------------
        veiculo = Veiculo.objects.create(
            contrato=contrato_obj,
            placa=placa,
            modelo=request.POST.get("modelo", "").strip(),
            marca=request.POST.get("marca", "").strip(),
            ano=request.POST.get("ano"),
            km_atual=request.POST.get("km_atual", 0).replace(".", ""),
            cor=request.POST.get("cor", "").strip(),
            categoria=request.POST.get("categoria"),
            status=request.POST.get("status", "Disponivel"),
            renavam=renavam,
            tipo=request.POST.get("tipo"),
            combustivel=request.POST.get("combustivel"),
            tipo_propriedade=request.POST.get("tipo_propriedade"),
            tag_interna=tag_interna,
            #tag_cliente=tag_cliente,
            licenciamento_vencimento=request.POST.get("licenciamento_vencimento") or None,
            seguro=request.POST.get("seguro") == "True",
            seguro_validade=request.POST.get("seguro_validade") or None,
            apolice_numero=apolice,
            observacoes=request.POST.get("observacoes", "").strip(),
            ativo=True,
        )

        messages.success(request, f"Veículo {placa} cadastrado com sucesso!")
        return redirect("detalhes_veiculo", id=veiculo.id)

    # GET → carregar form com contratos
    return render(request, "veiculos/novo_veiculo.html", {
        "contratos": Contrato.objects.all().order_by("id"),
        # Adicionar as choices
        "tipos_choices": Veiculo.TIPO_VEICULO_CHOICES,
        "categorias_choices": Veiculo.CATEGORIA_CHOICES,
        "combustiveis_choices": Veiculo.COMBUSTIVEL_CHOICES,
        "tipo_propriedade_choices": Veiculo.TIPO_PROPRIEDADE_CHOICES,
        "status_choices": Veiculo.STATUS_CHOICES,
    })





# EDITAR VEÍCULO (com validações completas + AJAX)
def editar_veiculo(request, id):
    veiculo = get_object_or_404(Veiculo, id=id)

    # Função para respostas AJAX
    def ajax_response(success, message, errors=None, redirect_url=None):
        response_data = {
            'success': success,
            'message': message,
        }
        if errors:
            response_data['errors'] = errors
        if redirect_url:
            response_data['redirect_url'] = redirect_url
        return JsonResponse(response_data, status=200 if success else 400)

    # Função auxiliar para retornar o form com contratos sempre carregados
    def render_form_com_contratos(msg=None):
        if msg:
            messages.error(request, msg)

        return render(request, "veiculos/editar.html", {
            "veiculo": veiculo,
            "contratos": Contrato.objects.all().order_by("id"),
            # ========== ADICIONAR AS CHOICES DO MODELO ==========
            "tipos_choices": Veiculo.TIPO_VEICULO_CHOICES,
            "categorias_choices": Veiculo.CATEGORIA_CHOICES,
            "combustiveis_choices": Veiculo.COMBUSTIVEL_CHOICES,
            "tipo_propriedade_choices": Veiculo.TIPO_PROPRIEDADE_CHOICES,
            "status_choices": Veiculo.STATUS_CHOICES,
            
        })

    if request.method == "POST":
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        # Normalização de campos
        placa = request.POST.get("placa", "").upper().strip()
        tag_interna = (request.POST.get("tag_interna") or "").strip()
        tag_cliente = (request.POST.get("tag_cliente") or "").strip()
        renavam = (request.POST.get("renavam") or "").strip()
        apolice = (request.POST.get("apolice_numero") or "").strip()
        ativo = request.POST.get("ativo") == "True"  # CORREÇÃO AQUI!

        errors = {}

        # ---------------------------------------------
        #  1) Validar PLACA
        # ---------------------------------------------
        if not placa:
            errors['placa'] = ["Placa é obrigatória!"]
        elif placa != veiculo.placa and Veiculo.objects.filter(placa=placa).exists():
            errors['placa'] = ["Já existe outro veículo com esta placa!"]

        # ---------------------------------------------
        #  2) Validar TAG INTERNA
        # ---------------------------------------------
        if tag_interna and tag_interna != veiculo.tag_interna:
            if Veiculo.objects.filter(tag_interna=tag_interna).exclude(id=id).exists():
                errors['tag_interna'] = ["Já existe outro veículo com esta TAG Interna!"]

            # ---------------------------------------------
        #  3) Validar TAG CLIENTE
        # ---------------------------------------------
        #if tag_cliente and tag_cliente != veiculo.tag_cliente:
           # if Veiculo.objects.filter(tag_cliente=tag_cliente).exclude(id=id).exists():
               #errors['tag_cliente'] = ["Já existe outro veículo com esta TAG do Cliente!"]


        # ---------------------------------------------
        #  4) Validar RENAVAM
        # ---------------------------------------------
        if renavam and renavam != veiculo.renavam:
            if Veiculo.objects.filter(renavam=renavam).exclude(id=id).exists():
                errors['renavam'] = ["Já existe outro veículo com este Renavam!"]

        # ---------------------------------------------
        #  5) Validar APÓLICE
        # ---------------------------------------------
        if apolice and apolice != veiculo.apolice_numero:
            if Veiculo.objects.filter(apolice_numero=apolice).exclude(id=id).exists():
                errors['apolice_numero'] = ["Este número de apólice já está vinculado a outro veículo!"]

        # ---------------------------------------------
        #  6) Campos obrigatórios básicos
        # ---------------------------------------------
        if not request.POST.get("modelo"):
            errors['modelo'] = ["Modelo é obrigatório!"]
        if not request.POST.get("marca"):
            errors['marca'] = ["Marca é obrigatória!"]
        if not request.POST.get("ano"):
            errors['ano'] = ["Ano é obrigatório!"]

        # Se houver erros, retorna
        if errors:
            if is_ajax:
                return ajax_response(False, "Por favor, corrija os erros abaixo.", errors)
            else:
                # Para requests não-AJAX, junta todos os erros
                error_messages = []
                for field_msgs in errors.values():
                    error_messages.extend(field_msgs)
                return render_form_com_contratos("<br>".join(error_messages))

        # ---------------------------------------------
        #  7) Atualizar CONTRATO
        # ---------------------------------------------
        contrato_id = request.POST.get("id_contrato")
        contrato_obj = Contrato.objects.filter(id=contrato_id).first() if contrato_id else None

        # ---------------------------------------------
        #  8) Atualização dos demais campos
        # ---------------------------------------------
        try:
            # Campos obrigatórios
            veiculo.contrato = contrato_obj
            veiculo.placa = placa
            veiculo.modelo = request.POST.get("modelo", "").strip()
            veiculo.marca = request.POST.get("marca", "").strip()
            veiculo.ano = request.POST.get("ano")
            
            # KM (remove pontos da formatação)
            km_atual_str = request.POST.get("km_atual", "0").replace(".", "")
            veiculo.km_atual = int(km_atual_str) if km_atual_str.isdigit() else 0
            
            # Demais campos
            veiculo.cor = request.POST.get("cor", "").strip()
            veiculo.categoria = request.POST.get("categoria")
            veiculo.status = request.POST.get("status", "Disponivel")
            veiculo.renavam = renavam
            veiculo.tipo = request.POST.get("tipo")
            veiculo.combustivel = request.POST.get("combustivel")
            veiculo.tipo_propriedade = request.POST.get("tipo_propriedade")
            veiculo.tag_interna = tag_interna
            veiculo.tag_cliente = tag_cliente
            veiculo.licenciamento_vencimento = request.POST.get("licenciamento_vencimento") or None
            veiculo.seguro = request.POST.get("seguro") == "True"
            veiculo.seguro_validade = request.POST.get("seguro_validade") or None
            veiculo.apolice_numero = apolice
            veiculo.observacoes = request.POST.get("observacoes", "").strip()
            veiculo.ativo = ativo  # CORREÇÃO: usa o valor do formulário!

            veiculo.save()

            # Resposta de sucesso
            success_message = f"Veículo {veiculo.placa} atualizado com sucesso!"
            
            if is_ajax:
                return ajax_response(
                    True, 
                    success_message, 
                    redirect_url=reverse("detalhes_veiculo", args=[veiculo.id])
                )
            else:
                messages.success(request, success_message)
                return redirect("detalhes_veiculo", id=veiculo.id)

        except Exception as e:
            error_msg = f"Erro ao salvar veículo: {str(e)}"
            if is_ajax:
                return ajax_response(False, error_msg)
            else:
                return render_form_com_contratos(error_msg)

    # GET → tela de edição
    return render_form_com_contratos()



# EXCLUIR VEÍCULO, (APENAS INATIVA)
def excluir_veiculo(request, id):
    veiculo = get_object_or_404(Veiculo, id=id)
    
    if request.method == "POST":
        placa = veiculo.placa
        veiculo.delete()
        messages.success(request, f"Veículo {placa} excluído com sucesso!")
        return redirect("lista_veiculos")
    
    return render(request, "veiculos/confirmar_exclusao.html", {"veiculo": veiculo})





# DETALHES DO VEÍCULO
def detalhes_veiculo(request, id):
    veiculo = get_object_or_404(Veiculo, id=id)
    
    # Histórico de movimentações (se houver modelo relacionado)
    historico_movimentacoes = []
    try:
        from movimentacoes.models import Movimentacao
        historico_movimentacoes = Movimentacao.objects.filter(
            veiculo=veiculo
        ).order_by('-data_saida')[:10]
    except:
        pass
    
    context = {
        'veiculo': veiculo,
        'historico_movimentacoes': historico_movimentacoes,
    }
    
    return render(request, "veiculos/detalhes_veiculo.html", context)


# EXPORTAR VEÍCULOS PARA EXCEL
def exportar_excel(request):
    perfil, created = PerfilUsuario.objects.get_or_create(
        user=request.user,
        defaults={"contrato": "Geral", "nivel": "basico"}
    )

    # Filtros recebidos via GET
    status_filter = request.GET.get("status", "").strip()
    search_query = request.GET.get("search", "").strip()
    categoria_filter = request.GET.get("categoria", "").strip()
    localizacao_filter = request.GET.get("localizacao", "").strip()
    contrato_filter = request.GET.get("contrato", "").strip()

    # Base - exportar tanto ativos quanto inativos se solicitado
    veiculos = Veiculo.objects.all().order_by("-ativo", "placa")

    # Permissão por contrato (não ADM)
    if perfil.nivel != "adm" and perfil.contrato:
        veiculos = veiculos.filter(contrato=perfil.contrato)

    # Filtro status
    if status_filter:
        mapa = {
            "disponivel": "Disponivel",
            "manutencao": "Manutencao",
        }
        
        if status_filter == "inativo":
            veiculos = veiculos.filter(ativo=False)
        elif status_filter in mapa:
            veiculos = veiculos.filter(ativo=True, status=mapa[status_filter])
        elif status_filter == "indisponivel":
            veiculos = veiculos.filter(ativo=True).exclude(status="Disponivel").exclude(status="Manutencao")
        else:
            if status_filter in dict(Veiculo.STATUS_CHOICES):
                veiculos = veiculos.filter(ativo=True, status=status_filter)

    # Filtro de busca
    if search_query:
        veiculos = veiculos.filter(
            Q(placa__icontains=search_query) |
            Q(modelo__icontains=search_query) |
            Q(marca__icontains=search_query) |
            Q(renavam__icontains=search_query) |
            Q(tag_interna__icontains=search_query) |
            Q(tag_cliente__icontains=search_query) |
            Q(contrato__icontains=search_query)
        )

    # Filtro categoria
    if categoria_filter:
        veiculos = veiculos.filter(categoria=categoria_filter)

    # Filtro localização (se existir no model)
    if localizacao_filter and hasattr(Veiculo, "localizacao"):
        veiculos = veiculos.filter(localizacao__icontains=localizacao_filter)

    # Filtro contrato
    if contrato_filter:
        veiculos = veiculos.filter(contrato__iexact=contrato_filter)

    # Criar workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Veículos"

    # Estilo cabeçalho
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="00594C", end_color="00594C", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "Ativo",
        "Placa",
        "Contrato",
        "Tag Interna",
        "Tag Cliente",
        "Marca",
        "Modelo",
        "Ano",
        "Cor",
        "Categoria",
        "Tipo",
        "Combustível",
        "KM Atual",
        "Propriedade",
        "Status",
        "Renavam",
        "IPVA Venc.",
        "Licenciamento Venc.",
        "Seguro Ativo",
        "Validade Seguro",
        "Apólice",
        "Observações",
    ]

    # Escrever cabeçalhos
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        sheet.column_dimensions[cell.column_letter].width = max(len(header) + 2, 14)

    # Escrever dados
    for row, v in enumerate(veiculos, start=2):
        sheet.cell(row=row, column=1, value="Sim" if v.ativo else "Não")
        sheet.cell(row=row, column=2, value=v.placa)
        sheet.cell(row=row, column=3, value=str(v.contrato) if v.contrato else "")  # CONVERTE PARA STRING
        sheet.cell(row=row, column=4, value=v.tag_interna)
        sheet.cell(row=row, column=5, value=v.tag_cliente)
        sheet.cell(row=row, column=6, value=v.marca)
        sheet.cell(row=row, column=7, value=v.modelo)
        sheet.cell(row=row, column=8, value=v.ano)
        sheet.cell(row=row, column=9, value=v.cor)
        sheet.cell(row=row, column=10, value=v.categoria)
        sheet.cell(row=row, column=11, value=v.tipo)
        sheet.cell(row=row, column=12, value=v.combustivel)
        sheet.cell(row=row, column=13, value=v.km_atual)
        sheet.cell(row=row, column=14, value=v.tipo_propriedade)
        sheet.cell(row=row, column=15, value=v.status)
        sheet.cell(row=row, column=16, value=v.renavam)
        sheet.cell(row=row, column=17, value=v.ipva_vencimento.strftime("%d/%m/%Y") if v.ipva_vencimento else "")
        sheet.cell(row=row, column=18, value=v.licenciamento_vencimento.strftime("%d/%m/%Y") if v.licenciamento_vencimento else "")
        sheet.cell(row=row, column=19, value="Sim" if v.seguro else "Não")
        sheet.cell(row=row, column=20, value=v.seguro_validade.strftime("%d/%m/%Y") if v.seguro_validade else "")
        sheet.cell(row=row, column=21, value=v.apolice_numero)
        sheet.cell(row=row, column=22, value=v.observacoes)

    # CORREÇÃO AQUI: Remover save_virtual_workbook e usar BytesIO
    from io import BytesIO
    
    # Criar buffer em memória
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)  # Voltar para o início do buffer
    
    # Adicionar timestamp ao nome do arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"veiculos_{timestamp}.xlsx"
    
    # Responder com arquivo
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# VALIDAÇÃO FLEXÍVEL DE PLACAS (MERCOSUL, ANTIGAS, EQUIPAMENTOS)
def validar_placa_flexivel(placa_str):
    """
    Valida diferentes formatos de placa:
    - Formato padrão Mercosul: ABC1D23
    - Formato antigo: ABC-1234
    - Formato especial para equipamentos: DTWA01064 (letras + números)
    """
    placa_upper = placa_str.upper().strip()
    
    # Remover hífens e espaços
    placa_clean = placa_upper.replace("-", "").replace(" ", "")
    
    # 1. Verificar se é placa especial (equipamentos/grupos geradores)
    # Padrão: 3-4 letras + 4-5 números
    if re.match(r'^[A-Z]{3,4}\d{4,5}$', placa_clean):
        return True, "Placa de equipamento válida", "equipamento"
    
    # 2. Formato Mercosul (ABC1D23)
    if re.match(r'^[A-Z]{3}\d[A-Z]\d{2}$', placa_clean):
        return True, "Placa Mercosul válida", "mercosul"
    
    # 3. Formato antigo com hífen (ABC-1234)
    if re.match(r'^[A-Z]{3}-\d{4}$', placa_upper):
        return True, "Placa antiga válida", "antiga"
    
    # 4. Formato antigo sem hífen (ABC1234)
    if re.match(r'^[A-Z]{3}\d{4}$', placa_clean):
        return True, "Placa antiga válida", "antiga"
    
    return False, f"Placa '{placa_str}' inválida. Formatos aceitos: ABC-1234, ABC1234, ABC1D23 ou DTWA01064 (equipamentos)", None

def formatar_placa_flexivel(placa_str):
    """Formata a placa mantendo o formato original do equipamento"""
    placa_upper = placa_str.upper().strip()
    
    # Se já está no formato especial, retorna como está
    if re.match(r'^[A-Z]{3,4}\d{4,5}$', placa_upper.replace("-", "")):
        return placa_upper.replace("-", "")
    
    # Formatar placa antiga com hífen
    if re.match(r'^[A-Z]{3}\d{4}$', placa_upper.replace("-", "")):
        placa_sem_hifen = placa_upper.replace("-", "")
        return f"{placa_sem_hifen[:3]}-{placa_sem_hifen[3:]}"
    
    # Para outros formatos, retorna sem formatação especial
    return placa_upper


# VIEW PARA ATUALIZAR KM DO VEÍCULO
def atualizar_km(request, id):
    veiculo = get_object_or_404(Veiculo, id=id)
    
    if request.method == "POST":
        try:
            novo_km = request.POST.get("novo_km")
            if novo_km:
                veiculo.km_atual = novo_km
                veiculo.save()
                messages.success(request, f"KM atualizado para {novo_km} km")
        except Exception as e:
            messages.error(request, f"Erro ao atualizar KM: {str(e)}")
    
    return redirect("detalhes_veiculo", id=id)


# VIEW PARA ALTERAR STATUS DO VEÍCULO
def alterar_status(request, id):
    veiculo = get_object_or_404(Veiculo, id=id)
    
    if request.method == "POST":
        try:
            novo_status = request.POST.get("novo_status")
            if novo_status in dict(Veiculo.STATUS_CHOICES).keys():
                veiculo.status = novo_status
                veiculo.save()
                messages.success(request, f"Status alterado para {veiculo.get_status_display()}")
        except Exception as e:
            messages.error(request, f"Erro ao alterar status: {str(e)}")
    
    return redirect("detalhes_veiculo", id=id)





# ========== FUNÇÕES DE NORMALIZAÇÃO ==========

def normalizar_tipo(tipo_raw):
    """Normaliza o tipo para um valor válido do modelo"""
    if not tipo_raw or pd.isna(tipo_raw):
        return "Equipamento"
    
    tipo_lower = str(tipo_raw).lower().strip()
    
    # Mapeamento de valores comuns
    mapeamento = {
        'veículo': 'Veiculo',
        'veiculos': 'Veiculo',
        'van': 'Van',
        'ônibus': 'Onibus',
        'onibus': 'Onibus',
        'micro-ônibus': 'Onibus',
        'caminhao': 'Caminhao',
        'caminhão': 'Caminhao',
        'utilitario': 'Utilitario',
        'equipamento': 'Equipamento',
        'implemento': 'Implemento',
        'reboque': 'Reboque',
        'carro': 'Carro',
    }
    
    for key, value in mapeamento.items():
        if key in tipo_lower:
            return value
    
    return "Equipamento"

def normalizar_categoria(categoria_raw):
    """Normaliza a categoria para um valor válido do modelo"""
    if not categoria_raw or pd.isna(categoria_raw):
        return "Outros"
    
    categoria_lower = str(categoria_raw).lower().strip()
    
    mapeamento = {
        'leve': 'Leve',
        'medio': 'Medio',
        'pesado': 'Pesado',
        'equipamento': 'Equipamento',
        'implemento': 'Implemento',
        'van': 'VAN',
        'basculante': 'Pesado',
        'transporte': 'Pesado',
        'pipa': 'Pesado',
        'lubrificante': 'Pesado',
        'apoio': 'Medio',
        'gerador': 'Equipamento',
        'escavadeira': 'Equipamento',
        'pá mecânica': 'Equipamento',
    }
    
    for key, value in mapeamento.items():
        if key in categoria_lower:
            return value
    
    return "Outros"

def normalizar_combustivel(combustivel_raw):
    """Normaliza o combustível para um valor válido do modelo"""
    if not combustivel_raw or pd.isna(combustivel_raw):
        return "N/A"
    
    combustivel_lower = str(combustivel_raw).lower().strip()
    
    if 'diesel' in combustivel_lower:
        return "Diesel"
    elif 'flex' in combustivel_lower:
        return "Flex"
    elif 'gasolina' in combustivel_lower:
        return "Gasolina"
    elif 'etanol' in combustivel_lower or 'álcool' in combustivel_lower:
        return "Etanol"
    elif 'eletrico' in combustivel_lower:
        return "Eletrico"
    else:
        return "N/A"

def normalizar_propriedade(propriedade_raw):
    """Normaliza o tipo de propriedade para um valor válido do modelo"""
    if not propriedade_raw or pd.isna(propriedade_raw):
        return "Proprio"
    
    propriedade_lower = str(propriedade_raw).lower().strip()
    
    if 'proprio' in propriedade_lower or 'próprio' in propriedade_lower:
        return "Proprio"
    elif 'locado' in propriedade_lower:
        return "Locado"
    elif 'comodato' in propriedade_lower:
        return "Comodato"
    elif 'arrendado' in propriedade_lower:
        return "Arrendado"
    else:
        return "Proprio"

def normalizar_status(status_raw):
    """Normaliza o status para um valor válido do modelo"""
    if not status_raw or pd.isna(status_raw):
        return "Disponivel"
    
    status_lower = str(status_raw).lower().strip()
    
    if 'disponivel' in status_lower or 'disponível' in status_lower:
        return "Disponivel"
    elif 'transito' in status_lower:
        return "EmTransito"
    elif 'manutencao' in status_lower:
        return "Manutencao"
    elif 'reservado' in status_lower:
        return "Reservado"
    else:
        return "Disponivel"

# ========== VALIDAÇÃO MÍNIMA ==========

def validar_importacao_veiculos(df, contrato):
    """Validação mínima - apenas verifica se tem dados básicos"""
    resultado = {
        "erros": [],
        "linhas_validas": 0,
        "linhas_invalidas": 0
    }
    
    # Limpar nomes das colunas
    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    
    # Verificar colunas obrigatórias
    colunas_obrigatorias = ['placa', 'marca', 'modelo']
    colunas_faltando = [col for col in colunas_obrigatorias if col not in df.columns]
    
    if colunas_faltando:
        resultado["erros"].append({
            "linha": "CABEÇALHO",
            "campo": ", ".join(colunas_faltando),
            "erro": f"Colunas obrigatórias faltando: {', '.join(colunas_faltando)}"
        })
        return resultado
    
    for idx, row in df.iterrows():
        numero_linha = idx + 2
        erros_linha = []
        
        # Verificar marca
        marca = str(row.get("marca", "")).strip() if pd.notna(row.get("marca")) else ""
        if not marca:
            erros_linha.append("Marca é obrigatória")
        
        # Verificar modelo
        modelo = str(row.get("modelo", "")).strip() if pd.notna(row.get("modelo")) else ""
        if not modelo:
            erros_linha.append("Modelo é obrigatório")
        
        if erros_linha:
            resultado["linhas_invalidas"] += 1
            placa = str(row.get("placa", "")).upper().strip() if pd.notna(row.get("placa")) else "N/A"
            resultado["erros"].append({
                "linha": numero_linha,
                "placa": placa if placa else "N/A",
                "erros": erros_linha
            })
        else:
            resultado["linhas_validas"] += 1
    
    return resultado

# ========== PROCESSAMENTO PRINCIPAL ==========

def processar_importacao_veiculos_validada(df, contrato, usuario):
    """Processa a importação - ACEITA QUALQUER VALOR e normaliza"""
    
    resultados = {
        "criados": 0,
        "atualizados": 0,
        "total": 0,
        "erros": []
    }
    
    # Limpar nomes das colunas
    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    
    for idx, row in df.iterrows():
        try:
            # ========== PLACA / IDENTIFICADOR ==========
            placa_raw = str(row.get("placa", "")).upper().strip() if pd.notna(row.get("placa")) else ""
            
            # Se não tem placa ou é NAN, gera um identificador único
            if not placa_raw or placa_raw in ["NAN", "N/A", "NAN/N/A", ""]:
                tag = str(row.get("tag_interna", "")).upper().strip() if pd.notna(row.get("tag_interna")) else ""
                if tag and tag not in ["NAN", "N/A", ""]:
                    placa = tag
                else:
                    placa = f"EQ_{contrato.id}_{idx+2}_{usuario.id}"
            else:
                placa = placa_raw
            
            # ========== RENAVAM ==========
            renavam = None
            if pd.notna(row.get("renavam")) and str(row.get("renavam")).strip():
                renavam_raw = str(row.get("renavam")).strip()
                renavam = ''.join(filter(str.isdigit, renavam_raw)) if renavam_raw else None
            
            # ========== TAG INTERNA ==========
            tag_interna = None
            if pd.notna(row.get("tag_interna")) and str(row.get("tag_interna")).strip():
                tag_interna = str(row.get("tag_interna")).upper().strip()
                tag_interna = tag_interna.replace("-", "").replace(" ", "")
            
            # ========== DADOS BÁSICOS ==========
            marca = str(row.get("marca", "")).strip() if pd.notna(row.get("marca")) else ""
            modelo = str(row.get("modelo", "")).strip() if pd.notna(row.get("modelo")) else ""
            
            # ========== ANO ==========
            ano = 0
            if pd.notna(row.get("ano")):
                try:
                    ano = int(float(row.get("ano")))
                except:
                    ano = 0
            
            # ========== COR ==========
            cor = str(row.get("cor")) if pd.notna(row.get("cor")) else ""
            
            # ========== KM ATUAL ==========
            km_atual = 0
            if pd.notna(row.get("km_atual")):
                try:
                    km_atual = int(float(row.get("km_atual")))
                except:
                    km_atual = 0
            
            # ========== NORMALIZAR CAMPOS COM CHOICES ==========
            tipo_raw = str(row.get("tipo")) if pd.notna(row.get("tipo")) else "Equipamento"
            tipo = normalizar_tipo(tipo_raw)
            
            categoria_raw = str(row.get("categoria")) if pd.notna(row.get("categoria")) else "Outros"
            categoria = normalizar_categoria(categoria_raw)
            
            combustivel_raw = str(row.get("combustivel")) if pd.notna(row.get("combustivel")) else "N/A"
            combustivel = normalizar_combustivel(combustivel_raw)
            
            tipo_propriedade_raw = str(row.get("tipo_propriedade")) if pd.notna(row.get("tipo_propriedade")) else "Proprio"
            tipo_propriedade = normalizar_propriedade(tipo_propriedade_raw)
            
            status_raw = str(row.get("status")) if pd.notna(row.get("status")) else "Disponivel"
            status = normalizar_status(status_raw)
            
            # ========== TAG CLIENTE ==========
            tag_cliente = ""
            if pd.notna(row.get("tag_cliente")):
                tag_cliente = str(row.get("tag_cliente")).upper().strip()
            
            # ========== DATAS ==========
            ipva_vencimento = None
            if pd.notna(row.get("ipva_vencimento")):
                try:
                    ipva_vencimento = pd.to_datetime(row.get("ipva_vencimento")).date()
                except:
                    pass
            
            licenciamento_vencimento = None
            if pd.notna(row.get("licenciamento_vencimento")):
                try:
                    licenciamento_vencimento = pd.to_datetime(row.get("licenciamento_vencimento")).date()
                except:
                    pass
            
            seguro_validade = None
            if pd.notna(row.get("seguro_validade")):
                try:
                    seguro_validade = pd.to_datetime(row.get("seguro_validade")).date()
                except:
                    pass
            
            # ========== SEGURO ==========
            seguro = False
            if pd.notna(row.get("seguro")):
                valor = str(row.get("seguro")).lower().strip()
                seguro = valor in ["sim", "true", "1", "ok", "ativo", "yes", "s", "x", "pago"]
            
            # ========== APOLICE E OBSERVAÇÕES ==========
            apolice_numero = str(row.get("apolice_numero")) if pd.notna(row.get("apolice_numero")) else ""
            observacoes = str(row.get("observacoes")) if pd.notna(row.get("observacoes")) else ""
            
            # ========== VERIFICAR SE JÁ EXISTE ==========
            veiculo_existente = Veiculo.objects.filter(placa=placa).first()
            
            if veiculo_existente:
                # ATUALIZAR
                km_anterior = veiculo_existente.km_atual
                
                veiculo_existente.renavam = renavam
                veiculo_existente.marca = marca
                veiculo_existente.modelo = modelo
                veiculo_existente.ano = ano
                veiculo_existente.cor = cor
                veiculo_existente.contrato = contrato
                veiculo_existente.tag_interna = tag_interna
                veiculo_existente.tag_cliente = tag_cliente
                veiculo_existente.tipo = tipo
                veiculo_existente.categoria = categoria
                veiculo_existente.combustivel = combustivel
                veiculo_existente.km_atual = km_atual
                veiculo_existente.tipo_propriedade = tipo_propriedade
                veiculo_existente.status = status
                veiculo_existente.ativo = True
                veiculo_existente.ipva_vencimento = ipva_vencimento
                veiculo_existente.licenciamento_vencimento = licenciamento_vencimento
                veiculo_existente.seguro = seguro
                veiculo_existente.seguro_validade = seguro_validade
                veiculo_existente.apolice_numero = apolice_numero
                veiculo_existente.observacoes = observacoes
                
                veiculo_existente.save()
                
                # Registrar histórico de KM
                if km_atual > 0 and km_anterior != km_atual:
                    HistoricoKM.objects.create(
                        veiculo=veiculo_existente,
                        km_anterior=km_anterior,
                        km_novo=km_atual,
                        origem="IMPORTACAO"
                    )
                
                resultados["atualizados"] += 1
            else:
                # CRIAR NOVO
                veiculo = Veiculo(
                    placa=placa,
                    renavam=renavam,
                    marca=marca,
                    modelo=modelo,
                    ano=ano,
                    cor=cor,
                    contrato=contrato,
                    tag_interna=tag_interna,
                    tag_cliente=tag_cliente,
                    tipo=tipo,
                    categoria=categoria,
                    combustivel=combustivel,
                    km_atual=km_atual,
                    tipo_propriedade=tipo_propriedade,
                    status=status,
                    ativo=True,
                    origem='IMPORTACAO',
                    ipva_vencimento=ipva_vencimento,
                    licenciamento_vencimento=licenciamento_vencimento,
                    seguro=seguro,
                    seguro_validade=seguro_validade,
                    apolice_numero=apolice_numero,
                    observacoes=observacoes
                )
                
                veiculo.save()
                
                # Registrar histórico de KM inicial
                if km_atual > 0:
                    HistoricoKM.objects.create(
                        veiculo=veiculo,
                        km_anterior=0,
                        km_novo=km_atual,
                        origem="IMPORTACAO"
                    )
                
                resultados["criados"] += 1
            
            resultados["total"] += 1
                
        except Exception as e:
            resultados["erros"].append(f"Linha {idx+2}: {str(e)}")
            import traceback
            traceback.print_exc()
    
    return resultados

# ========== VIEW PRINCIPAL ==========

@login_required
def importar_veiculos(request):
    """View para importar veículos via Excel com validação mínima"""
    
    perfil = request.user.perfilusuario
    
    if perfil.nivel != "adm":
        messages.error(request, "Apenas administradores podem importar dados.")
        return redirect("lista_veiculos")
    
    # Limpar sessão
    if request.method == "GET":
        for key in ['importacao_erros', 'importacao_linhas_validas', 'importacao_linhas_invalidas']:
            if key in request.session:
                del request.session[key]
    
    if request.method == "POST":
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        arquivo = request.FILES.get("arquivo")
        contrato_id = request.POST.get("contrato")
        
        if not arquivo:
            if is_ajax:
                return JsonResponse({'success': False, 'mensagem': 'Selecione um arquivo.'})
            messages.error(request, "Selecione um arquivo.")
            return redirect("importar_veiculos")
        
        if not contrato_id:
            if is_ajax:
                return JsonResponse({'success': False, 'mensagem': 'Selecione um contrato.'})
            messages.error(request, "Selecione um contrato.")
            return redirect("importar_veiculos")
        
        try:
            contrato = Contrato.objects.get(id=contrato_id)
        except Contrato.DoesNotExist:
            if is_ajax:
                return JsonResponse({'success': False, 'mensagem': 'Contrato não encontrado.'})
            messages.error(request, "Contrato não encontrado.")
            return redirect("importar_veiculos")
        
        # Verificar extensão
        if not arquivo.name.endswith(('.xlsx', '.xls')):
            if is_ajax:
                return JsonResponse({'success': False, 'mensagem': 'Use .xlsx ou .xls'})
            messages.error(request, "Formato inválido. Use .xlsx ou .xls")
            return redirect("importar_veiculos")
        
        # Ler arquivo
        try:
            df = pd.read_excel(arquivo)
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'mensagem': f'Erro: {str(e)}'})
            messages.error(request, f"Erro: {str(e)}")
            return redirect("importar_veiculos")
        
        if df.empty:
            if is_ajax:
                return JsonResponse({'success': False, 'mensagem': 'Planilha vazia.'})
            messages.error(request, "Planilha vazia.")
            return redirect("importar_veiculos")
        
        # Validação mínima
        validacao = validar_importacao_veiculos(df, contrato)
        
        if validacao["erros"]:
            request.session['importacao_erros'] = validacao["erros"]
            request.session['importacao_linhas_validas'] = validacao["linhas_validas"]
            request.session['importacao_linhas_invalidas'] = len(validacao["erros"])
            
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'mensagem': f'{len(validacao["erros"])} erro(s) encontrados.',
                    'erros': validacao["erros"],
                    'recarregar': True
                })
            messages.error(request, f"{len(validacao['erros'])} erro(s) encontrados.")
            return redirect("importar_veiculos")
        
        # Processar importação
        try:
            resultados = processar_importacao_veiculos_validada(df, contrato, request.user)
            
            # Limpar sessão
            for key in ['importacao_erros', 'importacao_linhas_validas', 'importacao_linhas_invalidas']:
                if key in request.session:
                    del request.session[key]
            
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'criados': resultados['criados'],
                    'atualizados': resultados['atualizados'],
                    'total': resultados['total']
                })
            
            messages.success(request, f"✅ {resultados['criados']} criados, {resultados['atualizados']} atualizados. Total: {resultados['total']}")
            return redirect("lista_veiculos")
            
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'mensagem': f'Erro: {str(e)}'})
            messages.error(request, f"Erro: {str(e)}")
            return redirect("importar_veiculos")
    
    # GET
    contratos = Contrato.objects.filter(ativo=True).order_by("nome")
    return render(request, "veiculos/importar.html", {
        "contratos": contratos,
    })


def baixar_modelo_veiculos(request):
    """Baixa modelo de planilha para veículos"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Veículos"
    
    # Cabeçalhos
    headers = [
        "placa", "renavam", "marca", "modelo", "ano", "cor",
        "tag_interna", "tag_cliente", "tipo", "categoria",
        "combustivel", "km_atual", "tipo_propriedade", "status",
        "ipva_vencimento", "licenciamento_vencimento", "seguro",
        "seguro_validade", "apolice_numero", "observacoes"
    ]
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="00594C", end_color="00594C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Dados de exemplo
    exemplos = [
        ["OTE-7884", "12345678901", "TOYOTA", "HILUX 3.0 4X4", 2023, "BRANCA",
         "VA-100", "", "Utilitario", "Leve", "Diesel", 75000, "Proprio", "Disponivel",
         "31/12/2025", "31/12/2025", "Sim", "31/12/2025", "AP-123456", ""],
        ["GAV-895", "98765432109", "FIAT", "TITANO 2.0", 2022, "PRATA",
         "VA-101", "", "Utilitario", "Leve", "Diesel", 45000, "Proprio", "Disponivel",
         "31/12/2025", "31/12/2025", "Sim", "31/12/2025", "AP-789012", ""],
        ["DFR-8425", "45678912345", "FORD", "RANGER 2.2 XLS", 2024, "PRETA",
         "VA-102", "", "Utilitario", "Leve", "Diesel", 12000, "Locado", "Disponivel",
         "31/12/2026", "31/12/2026", "Sim", "31/12/2026", "AP-345678", ""],
    ]
    
    for row_idx, exemplo in enumerate(exemplos, 2):
        for col_idx, valor in enumerate(exemplo, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="modelo_veiculos.xlsx"'
    wb.save(response)
    return response