from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Q
from datetime import date as dt_date, timedelta  # Renomeado para evitar conflito
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Motorista
from contratos.models import Contrato
from contas.models import PerfilUsuario
from datetime import date

import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from openpyxl import Workbook
from .models import Motorista
from contratos.models import Contrato


# ----------------------------------------------------------------------
# LISTA MOTORISTAS (com filtro por contrato do usuário)
# ----------------------------------------------------------------------
def lista_motoristas(request):
    perfil = request.user.perfilusuario  # PerfilUsuario com contrato FK

    status = request.GET.get("status") or "ativo"
    busca = request.GET.get("search")

    # Base
    motoristas = Motorista.objects.all()

    # 🔥 Filtrar por contrato se não for ADMIN
    if perfil.nivel != "adm" and perfil.contrato_id:
        motoristas = motoristas.filter(contrato_id=perfil.contrato_id)

    # Filtro por status
    if status == "ativo":
        motoristas = motoristas.filter(ativo=True)
    elif status == "inativo":
        motoristas = motoristas.filter(ativo=False)
    elif status == "vencimento":
        limite = dt_date.today() + timedelta(days=60)  # Usando dt_date em vez de date
        motoristas = motoristas.filter(
            cnh_vencimento__lte=limite,
            ativo=True
        )

    # Busca inteligente
    if busca:
        motoristas = motoristas.filter(
            Q(nome__icontains=busca) |
            Q(cpf__icontains=busca) |
            Q(telefone__icontains=busca) |
            Q(cnh_numero__icontains=busca) |
            Q(cnh_categoria__icontains=busca)
        )

    # Calcular estatísticas
    motoristas_ativos = Motorista.objects.filter(ativo=True)
    motoristas_inativos = Motorista.objects.filter(ativo=False)
    
    # Aplicar filtro de contrato nas estatísticas também
    if perfil.nivel != "adm" and perfil.contrato_id:
        motoristas_ativos = motoristas_ativos.filter(contrato_id=perfil.contrato_id)
        motoristas_inativos = motoristas_inativos.filter(contrato_id=perfil.contrato_id)
    
    motoristas_ativos_count = motoristas_ativos.count()
    motoristas_inativos_count = motoristas_inativos.count()
    
    # Calcular vencimentos (60 dias)
    limite_vencimento = dt_date.today() + timedelta(days=60)
    motoristas_vencendo = Motorista.objects.filter(
        cnh_vencimento__lte=limite_vencimento,
        ativo=True
    )
    
    # Aplicar filtro de contrato nos vencimentos também
    if perfil.nivel != "adm" and perfil.contrato_id:
        motoristas_vencendo = motoristas_vencendo.filter(contrato_id=perfil.contrato_id)
    
    motoristas_vencendo_count = motoristas_vencendo.count()

    return render(request, "motoristas/lista.html", {
        "motoristas": motoristas,
        "motoristas_ativos": motoristas_ativos_count,
        "motoristas_inativos": motoristas_inativos_count,
        "motoristas_vencendo": motoristas_vencendo_count,
        "status": status,
        "search": busca
    })



# CRIAR MOTORISTA (agora com combobox de contratos)
# ----------------------------------------------------------------------
def criar_motorista(request):
    # Buscar todos os contratos ativos para o combobox
    contratos = Contrato.objects.filter(ativo=True).order_by('nome')
    
    if request.method == "POST":
        # O campo agora é "contrato" (select) ao invés de "id_contrato"
        contrato_id = request.POST.get("contrato")
        contrato_obj = Contrato.objects.filter(id=contrato_id).first() if contrato_id else None
        
        # O campo ativo agora vem como "True"/"False" string
        ativo = request.POST.get("ativo") == "True"

        Motorista.objects.create(
            nome=request.POST["nome"],
            cpf=request.POST["cpf"],
            telefone=request.POST["telefone"],
            cnh_numero=request.POST["cnh_numero"],
            cnh_categoria=request.POST["cnh_categoria"],
            cnh_vencimento=request.POST["cnh_vencimento"],
            matricula=request.POST["matricula"],       
            ativo=ativo,
            contrato=contrato_obj
        )

        return redirect("lista_motoristas")

    return render(request, "motoristas/form.html", {
        "contratos": contratos,
        "motorista": None
    })


# ----------------------------------------------------------------------
# EDITAR MOTORISTA
# ----------------------------------------------------------------------
def editar_motorista(request, pk):
    motorista = get_object_or_404(Motorista, pk=pk)
    
    # Buscar todos os contratos ativos para o combobox
    contratos = Contrato.objects.filter(ativo=True).order_by('nome')

    if request.method == "POST":
        # Matrícula (tratar como números apenas)
        matricula = request.POST.get("matricula", "").strip()
        # Remover qualquer caractere não numérico
        matricula = ''.join(filter(str.isdigit, matricula))
        
        # Contrato
        contrato_id = request.POST.get("contrato")
        contrato_obj = Contrato.objects.filter(id=contrato_id).first() if contrato_id else None
        
        # Status ativo
        ativo = request.POST.get("ativo") == "True"

        # Atualizar motorista
        motorista.matricula = matricula
        motorista.nome = request.POST.get("nome", "").strip().upper()
        motorista.cpf = request.POST.get("cpf", "").strip()
        motorista.telefone = request.POST.get("telefone", "").strip()
        motorista.cnh_numero = request.POST.get("cnh_numero", "").strip()
        motorista.cnh_categoria = request.POST.get("cnh_categoria", "").strip().upper()
        
        # Tratar data de vencimento
        cnh_vencimento = request.POST.get("cnh_vencimento")
        if cnh_vencimento:
            try:
                from datetime import datetime
                motorista.cnh_vencimento = datetime.strptime(cnh_vencimento, "%Y-%m-%d").date()
            except:
                motorista.cnh_vencimento = None
        else:
            motorista.cnh_vencimento = None
        
        motorista.ativo = ativo
        motorista.contrato = contrato_obj

        motorista.save()
        return redirect("lista_motoristas")

    return render(request, "motoristas/form.html", {
        "motorista": motorista,
        "contratos": contratos
    })


# ----------------------------------------------------------------------
# EXCLUIR / DESATIVAR
# ----------------------------------------------------------------------
def excluir_motorista(request, pk):
    motorista = get_object_or_404(Motorista, pk=pk)
    motorista.ativo = False
    motorista.save()
    return redirect("lista_motoristas")


# ----------------------------------------------------------------------
# DETALHES
# ----------------------------------------------------------------------
def detalhes_motorista(request, pk):
    motorista = get_object_or_404(Motorista, pk=pk)
    
    # Calcular dias para vencimento na view
    dias_para_vencimento = None
    if motorista.cnh_vencimento:
        hoje = date.today()
        diferenca = motorista.cnh_vencimento - hoje
        dias_para_vencimento = diferenca.days
    
    context = {
        'motorista': motorista,
        'dias_para_vencimento': dias_para_vencimento,  # Calculado na view
    }
    
    return render(request, "motoristas/partials/detalhes.html", context)





# ----------------------------------------------------------------------
# ATIVAR / DESATIVAR
# ----------------------------------------------------------------------
def toggle_motorista(request, pk):
    motorista = get_object_or_404(Motorista, pk=pk)
    motorista.ativo = not motorista.ativo
    motorista.save()
    return redirect("lista_motoristas")


# ----------------------------------------------------------------------
# EXPORTAR MOTORISTAS
# ----------------------------------------------------------------------
def exportar_motoristas(request):
    status = request.GET.get("status")

    if status == "ativo":
        motoristas = Motorista.objects.filter(ativo=True)
    elif status == "inativo":
        motoristas = Motorista.objects.filter(ativo=False)
    elif status == "vencimento":
        limite = dt_date.today() + timedelta(days=60)  # Usando dt_date
        motoristas = Motorista.objects.filter(cnh_vencimento__lte=limite)
    else:
        motoristas = Motorista.objects.all()

    # EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Motoristas"

    headers = ["Nome", "CPF", "Telefone", "CNH Número", "Categoria", "Vencimento", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="00594C", end_color="00594C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Estilo
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

    # Linhas
    for m in motoristas:
        ws.append([
            m.nome,
            m.cpf,
            m.telefone,
            m.cnh_numero,
            m.cnh_categoria,
            m.cnh_vencimento.strftime("%d/%m/%Y") if m.cnh_vencimento else "",
            "Ativo" if m.ativo else "Inativo",
        ])

    # Zebra
    zebra_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
        if idx % 2 == 0:
            for cell in row:
                cell.fill = zebra_fill
                cell.border = thin_border

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=motoristas.xlsx"
    wb.save(response)

    return response


# ----------------------------------------------------------------------
# AJAX PARA FILTROS
# ----------------------------------------------------------------------
def buscar_motoristas_ajax(request):
    status = request.GET.get("status") or "ativo"
    busca = request.GET.get("search")
    perfil = request.user.perfilusuario

    motoristas = Motorista.objects.all()

    #  Filtro por contrato do usuário
    if perfil.nivel != "adm" and perfil.contrato_id:
        motoristas = motoristas.filter(contrato_id=perfil.contrato_id)

    # Filtros de status
    if status == "ativo":
        motoristas = motoristas.filter(ativo=True)
    elif status == "inativo":
        motoristas = motoristas.filter(ativo=False)
    elif status == "vencimento":
        limite = dt_date.today() + timedelta(days=60)  # Usando dt_date
        motoristas = motoristas.filter(cnh_vencimento__lte=limite, ativo=True)

    # Busca
    if busca:
        motoristas = motoristas.filter(
            Q(nome__icontains=busca) |
            Q(cpf__icontains=busca) |
            Q(telefone__icontains=busca) |
            Q(cnh_numero__icontains=busca) |
            Q(cnh_categoria__icontains=busca)
        )

    return render(request, "motoristas/partials/cards.html", {
        "motoristas": motoristas,
        "status": status
    })




# ==================== IMPORTAÇÃO DE MOTORISTAS ====================
@login_required
def importar_motoristas(request):
    """View para importar motoristas via Excel"""
    
    perfil = request.user.perfilusuario
    
    # Verificar se é ADM
    if perfil.nivel != "adm":
        messages.error(request, "Apenas administradores podem importar dados.")
        return redirect("lista_motoristas")
    
    if request.method == "POST":
        arquivo = request.FILES.get("arquivo")
        contrato_id = request.POST.get("contrato")
        
        if not arquivo:
            messages.error(request, "Selecione um arquivo para importar.")
            return redirect("importar_motoristas")
        
        try:
            contrato = Contrato.objects.get(id=contrato_id) if contrato_id else None
        except Contrato.DoesNotExist:
            messages.error(request, "Contrato não encontrado.")
            return redirect("importar_motoristas")
        
        # Ler o arquivo Excel
        try:
            df = pd.read_excel(arquivo)
        except Exception as e:
            messages.error(request, f"Erro ao ler o arquivo: {str(e)}")
            return redirect("importar_motoristas")
        
        # Processar importação
        resultados = processar_importacao_motoristas(df, contrato)
        
        # Mensagem de resultado
        if resultados["erros"]:
            messages.warning(
                request,
                f"Importação concluída: {resultados['criados']} criados, "
                f"{resultados['atualizados']} atualizados. "
                f"{len(resultados['erros'])} erros encontrados."
            )
            # Salvar erros na sessão para exibir depois
            request.session['importacao_erros'] = resultados["erros"][:20]
        else:
            messages.success(
                request,
                f"Importação concluída com sucesso! "
                f"{resultados['criados']} motoristas criados, "
                f"{resultados['atualizados']} atualizados."
            )
            # Limpar erros da sessão
            if 'importacao_erros' in request.session:
                del request.session['importacao_erros']
        
        return redirect("lista_motoristas")
    
    # GET - Mostrar formulário
    contratos = Contrato.objects.all().order_by("nome")
    return render(request, "motoristas/importar.html", {
        "contratos": contratos,
    })


def processar_importacao_motoristas(df, contrato):
    """Processa a importação dos motoristas"""
    
    resultados = {
        "criados": 0,
        "atualizados": 0,
        "erros": []
    }
    
    # Limpar nomes das colunas (remover espaços, converter para minúsculo)
    df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
    
    # Verificar se as colunas obrigatórias existem
    colunas_obrigatorias = ['nome', 'cpf', 'cnh_numero', 'cnh_categoria', 'matricula', 'cidade']
    colunas_faltando = [col for col in colunas_obrigatorias if col not in df.columns]
    
    if colunas_faltando:
        resultados["erros"].append(f"Colunas obrigatórias faltando: {', '.join(colunas_faltando)}")
        return resultados
    
    for idx, row in df.iterrows():
        try:
            # Validar campos obrigatórios
            nome = str(row.get("nome", "")).upper().strip()
            if not nome:
                resultados["erros"].append(f"Linha {idx+2}: Nome é obrigatório")
                continue
            
            cpf = str(row.get("cpf", "")).strip()
            if not cpf:
                resultados["erros"].append(f"Linha {idx+2}: CPF é obrigatório")
                continue
            
            # Limpar CPF (remover pontos, traços, barras e espaços)
            cpf_limpo = cpf.replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
            
            # Validar CPF tem 11 dígitos
            if len(cpf_limpo) != 11:
                resultados["erros"].append(f"Linha {idx+2}: CPF '{cpf}' inválido (deve ter 11 dígitos)")
                continue
            
            # CNH
            cnh_numero = str(row.get("cnh_numero", "")).strip()
            if not cnh_numero:
                resultados["erros"].append(f"Linha {idx+2}: Número da CNH é obrigatório")
                continue
            
            cnh_categoria = str(row.get("cnh_categoria", "")).upper().strip()
            if not cnh_categoria:
                resultados["erros"].append(f"Linha {idx+2}: Categoria da CNH é obrigatória")
                continue
            
            # Matrícula
            matricula = str(row.get("matricula", "")).strip()
            if not matricula:
                # Se não informada, gerar matrícula numérica baseada no CPF
                matricula = cpf_limpo[-6:]  # últimos 6 dígitos do CPF
            else:
                # Remover qualquer caractere não numérico (espaços, letras, etc)
                matricula = ''.join(filter(str.isdigit, matricula))
                if not matricula:
                    # Se ficou vazio, usar fallback
                    matricula = cpf_limpo[-6:]
            
            # Cidade
            cidade = str(row.get("cidade", "")).upper().strip()
            if not cidade:
                cidade = "PARAUAPEBAS"  # Valor padrão
            
            # Data de vencimento da CNH
            cnh_vencimento = None
            if pd.notna(row.get("cnh_vencimento")):
                try:
                    cnh_vencimento = pd.to_datetime(row.get("cnh_vencimento")).date()
                except:
                    pass
            
            # Telefone
            telefone = str(row.get("telefone", "")) if pd.notna(row.get("telefone")) else ""
            
            # Email
            email = str(row.get("email", "")) if pd.notna(row.get("email")) else ""
            
            # Endereço
            endereco = str(row.get("endereco", "")) if pd.notna(row.get("endereco")) else ""
            
            # Estado
            estado = str(row.get("estado", "")) if pd.notna(row.get("estado")) else ""
            
            # Buscar ou criar motorista (por CPF)
            motorista, created = Motorista.objects.update_or_create(
                cpf=cpf_limpo,
                defaults={
                    "nome": nome,
                    "cnh_numero": cnh_numero,
                    "cnh_categoria": cnh_categoria,
                    "cnh_vencimento": cnh_vencimento,
                    "matricula": matricula,
                    "telefone": telefone,
                    "email": email,
                    "endereco": endereco,
                    "cidade": cidade,
                    "estado": estado,
                    "contrato": contrato,
                    "ativo": True,
                }
            )
            
            if created:
                resultados["criados"] += 1
            else:
                resultados["atualizados"] += 1
                
        except Exception as e:
            resultados["erros"].append(f"Linha {idx+2}: {str(e)}")
    
    return resultados


# ==================== DOWNLOAD DO MODELO ====================
def baixar_modelo_motoristas(request):
    """Baixa modelo de planilha para motoristas"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Motoristas"
    
    # Cabeçalhos
    headers = [
        "nome", "cpf", "telefone", "cnh_numero", "cnh_categoria",
        "cnh_vencimento", "email", "matricula", "endereco", "cidade", "estado"
    ]
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="00594C", end_color="00594C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Dados de exemplo
    exemplos = [
        ["LAZARO COSTA SILVA", "123.456.789-00", "(94) 99999-9999", "12345678900", "D",
         "31/12/2028", "lazaro@email.com", "MAT-001", "RUA A, 123", "PARAUAPEBAS", "PA"],
        ["ANTONIO NUNES", "987.654.321-00", "(94) 88888-8888", "98765432100", "C",
         "30/06/2027", "antonio@email.com", "MAT-002", "RUA B, 456", "PARAUAPEBAS", "PA"],
        ["JANES SILVA", "456.789.123-00", "(94) 77777-7777", "45678912300", "D",
         "15/12/2029", "janes@email.com", "MAT-003", "RUA C, 789", "PARAUAPEBAS", "PA"],
    ]
    
    for row_idx, exemplo in enumerate(exemplos, 2):
        for col_idx, valor in enumerate(exemplo, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="modelo_motoristas.xlsx"'
    wb.save(response)
    return response    